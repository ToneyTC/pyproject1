from playwright.sync_api import sync_playwright
import os
import re
import time


def _选择excel文件():
    path = input("请输入 Excel 路径（留空则尝试弹窗选择）：").strip().strip('"')
    if path:
        return path
    try:
        import tkinter as tk
        from tkinter import filedialog

        root = tk.Tk()
        root.withdraw()
        file_path = filedialog.askopenfilename(
            title="选择 Excel 文件",
            filetypes=[("Excel files", "*.xlsx *.xlsm *.xltx *.xltm")],
        )
        root.destroy()
        return file_path
    except Exception:
        return ""


def _读取excel预览(file_path, max_rows=10):
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception as e:
        raise RuntimeError(f"未安装 openpyxl：{e}")

    if not os.path.exists(file_path):
        raise FileNotFoundError(f"文件不存在：{file_path}")

    wb = load_workbook(file_path, data_only=True, read_only=True)
    try:
        ws = wb.active
        rows = []
        for row in ws.iter_rows(values_only=True):
            vals = ["" if v is None else str(v).strip() for v in row]
            if any(vals):
                rows.append(vals)
            if len(rows) >= max_rows:
                break
        return ws.title, rows
    finally:
        wb.close()


def _读取excel数字代码列表(file_path, code_col=1):
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception as e:
        raise RuntimeError(f"未安装 openpyxl：{e}")

    if not os.path.exists(file_path):
        raise FileNotFoundError(f"文件不存在：{file_path}")

    wb = load_workbook(file_path, data_only=True, read_only=True)
    try:
        ws = wb.active
        codes = []
        idx = max(0, code_col - 1)
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or idx >= len(row):
                continue
            v = row[idx]
            if v is None:
                continue
            s = str(v).strip()
            if s.isdigit():
                codes.append(s)
        # 去重保序
        out = []
        seen = set()
        for c in codes:
            if c in seen:
                continue
            seen.add(c)
            out.append(c)
        return out
    finally:
        wb.close()


def 更新计划(page):
    print("已选择：更新计划")
    print(page.title())
    file_path = _选择excel文件()
    if not file_path:
        print("未选择 Excel 文件")
        return

    try:
        sheet_name, rows = _读取excel预览(file_path)
    except Exception as e:
        print(f"识别 Excel 失败：{e}")
        return

    print(f"已选择文件：{file_path}")
    print(f"工作表：{sheet_name}")
    if not rows:
        print("Excel 为空或没有可识别数据")
        return

    print("Excel 预览（前10行）：")
    for i, row in enumerate(rows, start=1):
        print(f"{i}: {row}")

    # 读取完整数字代码列表（默认第1列，第2行开始）
    try:
        codes = _读取excel数字代码列表(file_path, code_col=1)
    except Exception as e:
        print(f"读取 Excel 代码列表失败：{e}")
        return
    if not codes:
        print("⚠️ Excel 未读取到有效数字代码")
        return

    # 识别完成后：点击 生产管理 -> 工艺 -> 产品BOM
    page.click("b.quick-name:text('生产管理')", timeout=8000)
    page.click("span.title:text('工艺')", timeout=8000)
    page.click("div.child-menu-name:text('产品BOM')", timeout=8000)
    print("✅ 已进入：生产管理 -> 工艺 -> 产品BOM")
    # 进入 BOM 后先切换到「待发布」
    page.click("div.el-tabs__item:text('待发布')", timeout=8000)
    time.sleep(0.6)

    # 循环按 Excel 代码逐个查询：仅 1 行时若含半成品则查下一条，否则勾选并编辑
    bom_code_input = page.locator("input.el-input__inner[placeholder='请输入物料代码']").first

    def _行包含半成品或bcp(row):
        try:
            base = row.inner_text()
        except Exception:
            base = ""
        tips = ""
        try:
            tip_loc = row.locator(".tooltip-column-content")
            if tip_loc.count() > 0:
                tips = "\n".join([t.strip() for t in tip_loc.all_inner_texts() if t and t.strip()])
        except Exception:
            tips = ""
        txt = f"{base}\n{tips}"
        low = txt.lower()
        return ("半成品" in txt) or ("_bcp" in low) or ("bcp" in low)

    def _首行物料名称_按xpath():
        """
        优先使用你提供的绝对 XPath 读取第1行第5列（物料名称）。
        """
        xp = "/html/body/div[1]/div/div[2]/div/div[2]/div[2]/div/div/div/div/div[1]/div[1]/div[2]/div/div/div/div[3]/table/tbody/tr[1]/td[5]/div/div/div"
        loc = page.locator(f"xpath={xp}")
        if loc.count() == 0:
            return ""
        try:
            return loc.first.inner_text().strip()
        except Exception:
            return ""

    def _首行row_按xpath():
        """
        使用你提供的绝对 XPath 定位第一行 tr。
        """
        xp = "/html/body/div[1]/div/div[2]/div/div[2]/div[2]/div/div/div/div/div[1]/div[1]/div[2]/div/div/div/div[4]/div[2]/table/tbody/tr[1]"
        loc = page.locator(f"xpath={xp}")
        return loc.first if loc.count() > 0 else None

    # 产品 BOM 列表：第 tr 行第 1 列勾选框内核（label/span/span），与 #container 下表格一致
    _BOM列表勾选内核_XPATH = (
        '//*[@id="container"]/div/div/div/div[1]/div[1]/div[2]/div/div/div/div[4]/div[2]/table/tbody/tr[{n}]/td[1]/div/label/span/span'
    )

    def _按容器xpath点列表行勾选(n_tr_1based: int) -> bool:
        xp = _BOM列表勾选内核_XPATH.format(n=n_tr_1based)
        loc = page.locator(f"xpath={xp}")
        if loc.count() == 0:
            return False
        el = loc.first
        try:
            el.evaluate(
                "e => { e.scrollIntoView({ block: 'center', inline: 'nearest' }); e.click(); }"
            )
        except Exception:
            try:
                el.click(timeout=5000, force=True)
            except Exception:
                return False
        time.sleep(0.15)
        return True

    def _bom主表数据行():
        """
        仅主滚动表体的 tr。全页 `table.el-table__body tr` 会把左/右固定列各算一遍，1 条常显示为 3。
        """
        loc = page.locator(
            "#container .el-table__inner-wrapper > .el-table__body-wrapper tbody tr"
        )
        if loc.count() > 0:
            return loc
        return page.locator("#container .el-table__body-wrapper").first.locator("tbody tr")

    def _表格行滚入可视区(r):
        """el-table 常在 body-wrapper 内滚动，仅 scroll_into_view 易仍判为视口外。"""
        try:
            r.evaluate(
                """el => {
                let p = el;
                for (let i = 0; i < 12 && p; i++, p = p.parentElement) {
                  if (!p || !p.classList) continue;
                  if (p.classList.contains('el-table__body-wrapper')) {
                    const row = el.closest('tr');
                    if (row) {
                      const top = row.offsetTop - Math.max(0, (p.clientHeight - row.offsetHeight) / 2);
                      p.scrollTop = Math.max(0, top);
                    }
                    break;
                  }
                }
                el.scrollIntoView({ block: 'center', inline: 'nearest' });
            }"""
            )
        except Exception:
            try:
                r.scroll_into_view_if_needed(timeout=5000)
            except Exception:
                pass
        time.sleep(0.2)

    def _首列复选视觉上已勾选(first_cell) -> bool:
        """以 label.is-checked 为准；勿仅用原生 input.checked，否则易与 Element Plus 界面不同步却误判成功。"""
        try:
            return first_cell.locator("label.el-checkbox.is-checked").count() > 0
        except Exception:
            return False

    def _勾选表格行首列(row, list_row_index=None, use_bom_row_xpath=True) -> bool:
        """
        仅勾选表格行首列复选框（不点编辑）。
        list_row_index: 从 0 开始的行号；与 use_bom_row_xpath=True 时优先尝试 BOM 列表专用 XPath。
        工艺管理列表请传 use_bom_row_xpath=False，避免 XPath 点到别的表或无效节点。
        """
        _表格行滚入可视区(row)
        first_cell = row.locator("td").first
        selected = False

        if list_row_index is not None and use_bom_row_xpath:
            _按容器xpath点列表行勾选(list_row_index + 1)
            time.sleep(0.2)
            selected = _首列复选视觉上已勾选(first_cell)

        cb_input = first_cell.locator("input.el-checkbox__original").first
        cb_label = first_cell.locator("label.el-checkbox").first

        if not selected and cb_input.count() > 0:
            try:
                cb_input.evaluate(
                    """el => {
                    el.scrollIntoView({ block: 'center', inline: 'nearest' });
                    if (!el.checked) el.click();
                }"""
                )
                time.sleep(0.2)
                selected = _首列复选视觉上已勾选(first_cell)
            except Exception:
                selected = False

        if not selected and cb_label.count() > 0:
            try:
                cb_label.click(timeout=5000, force=True)
                selected = _首列复选视觉上已勾选(first_cell)
            except Exception:
                try:
                    cb_label.evaluate(
                        """el => {
                        el.scrollIntoView({ block: 'center', inline: 'nearest' });
                        el.click();
                    }"""
                    )
                    time.sleep(0.2)
                    selected = _首列复选视觉上已勾选(first_cell)
                except Exception:
                    selected = False

        if not selected:
            cb_inner = first_cell.locator("span.el-checkbox__inner").first
            if cb_inner.count() > 0:
                try:
                    cb_inner.click(timeout=5000, force=True)
                    selected = _首列复选视觉上已勾选(first_cell)
                except Exception:
                    try:
                        cb_inner.evaluate("el => { el.scrollIntoView({block:'center'}); el.click(); }")
                        time.sleep(0.2)
                        selected = _首列复选视觉上已勾选(first_cell)
                    except Exception:
                        pass

        if not selected and cb_input.count() > 0:
            try:
                cb_input.evaluate("el => { if (!el.checked) el.click(); }")
                time.sleep(0.25)
            except Exception:
                pass
            selected = _首列复选视觉上已勾选(first_cell)

        if not _首列复选视觉上已勾选(first_cell):
            for _ in range(15):
                time.sleep(0.12)
                if _首列复选视觉上已勾选(first_cell):
                    break

        if _首列复选视觉上已勾选(first_cell):
            print("✅ 行已勾选")
            return True

        print("⚠️ 勾选失败：未识别到已选中状态（以界面勾选样式为准）")
        return False

    def _勾选并点击编辑(row, list_row_index=None):
        """
        list_row_index: 查询结果表格中从 0 开始的行号；若提供则优先用 #container 下勾选 XPath 点击。
        """
        if not _勾选表格行首列(row, list_row_index):
            return False
        _表格行滚入可视区(row)
        # 点击编辑（行内优先）
        op_cell = row.locator("td").last
        btn = op_cell.locator("button:has-text('编辑'), a:has-text('编辑'), span:has-text('编辑'), [title*='编辑'], [aria-label*='编辑']").first
        if btn.count() > 0:
            try:
                btn.click(timeout=8000, force=True)
                return True
            except Exception:
                try:
                    btn.evaluate("el => { el.scrollIntoView({block:'center'}); el.click(); }")
                    return True
                except Exception:
                    pass
        # 兜底：工具栏编辑
        for tb_sel in (
            "button:has-text('编辑修改')",
            "button:has-text('编辑')",
        ):
            tb = page.locator(tb_sel).first
            if tb.count() > 0:
                try:
                    tb.click(timeout=8000, force=True)
                    return True
                except Exception:
                    try:
                        tb.evaluate("el => el.click()")
                        return True
                    except Exception:
                        pass
        return False

    def _进入编辑后检查投料工位与主料():
        """
        进入 BOM 编辑后：
        - 成品/半成品行将「是否主料」设为「是」
        - 投料工位未选时：点击后于「添加工位」弹窗搜索「机包」并保存
        - 结束后校验成品/半成品是否均为「主料=是」
        """
        time.sleep(1.5)
        rows = page.locator("table.el-table__body tr")
        row_count = rows.count()
        if row_count == 0:
            print("⚠️ 编辑页未找到物料表格行")
            return

        feed_ph = "input[placeholder*='投料工位']"

        def _resolve_feed_cell(r, cells, cell_count):
            feed_td = r.locator(f"td:has({feed_ph})").first
            if feed_td.count() > 0:
                return feed_td
            if cell_count >= 6:
                return cells.nth(cell_count - 3)
            return cells.nth(cell_count - 1)

        def _投料工位弹窗选机包并保存(feed_cell):
            def _wait_dialog_search_input(timeout_ms):
                """不依赖标题文案：弹窗标题可能是「添加工位」「选择工位」等，以搜索框为准。"""
                deadline = time.time() + timeout_ms / 1000.0
                while time.time() < deadline:
                    for ph_sel in (
                        "input[placeholder='请输入工位']",
                        "input[placeholder*='请输入'][placeholder*='工位']",
                        "input[placeholder*='工位']",
                    ):
                        loc = page.locator(f".el-dialog:visible {ph_sel}").first
                        if loc.count() > 0:
                            try:
                                loc.wait_for(state="visible", timeout=800)
                                return True
                            except Exception:
                                pass
                    time.sleep(0.12)
                return False

            opened = False
            click_targets = [
                feed_cell.locator(feed_ph).first,
                feed_cell.locator("input.el-input__inner").first,
                feed_cell.locator(".el-input__suffix .el-input__suffix-inner").first,
                feed_cell.locator(".el-input").first,
                feed_cell,
            ]
            for _round in range(4):
                for target in click_targets:
                    if target.count() == 0:
                        continue
                    try:
                        target.scroll_into_view_if_needed(timeout=4000)
                    except Exception:
                        pass
                    try:
                        target.click(timeout=5000, force=True)
                    except Exception:
                        try:
                            target.evaluate("el => el.click()")
                        except Exception:
                            continue
                    time.sleep(0.35)
                    if _wait_dialog_search_input(2800):
                        opened = True
                        break
                if opened:
                    break
                time.sleep(0.25)

            if not opened:
                print("⚠️ 点击投料工位后未出现工位搜索弹窗（未识别到「请输入工位」等输入框）")
                return False

            dlg = page.locator(".el-dialog:visible").filter(
                has=page.locator("input[placeholder='请输入工位']")
            ).first
            if dlg.count() == 0:
                dlg = page.locator(".el-dialog:visible").filter(
                    has=page.locator("input[placeholder*='工位']")
                ).first
            if dlg.count() == 0:
                dlg = page.locator(".el-dialog:visible").first
            try:
                dlg.wait_for(state="visible", timeout=5000)
            except Exception:
                pass

            q_in = dlg.locator("input[placeholder='请输入工位']").first
            if q_in.count() == 0:
                q_in = dlg.locator("input[placeholder*='工位']").first
            if q_in.count() > 0:
                q_in.fill("机包")
            q_btn = dlg.locator("button:has-text('查询')").first
            if q_btn.count() > 0:
                q_btn.click(timeout=5000)
                time.sleep(0.85)
            data_row = dlg.locator(".el-dialog__body table.el-table__body tbody tr").filter(has_text="机包").first
            if data_row.count() == 0:
                data_row = dlg.locator("table tbody tr").filter(has_text="机包").first
            if data_row.count() == 0:
                print("⚠️ 工位弹窗中未找到含「机包」的行")
                cxl = dlg.locator("button:has-text('取消')").first
                if cxl.count() > 0:
                    cxl.click(timeout=3000)
                time.sleep(0.3)
                return False
            cb = data_row.locator("label.el-checkbox").first
            if cb.count() > 0:
                cb.click(timeout=5000, force=True)
            else:
                data_row.click(timeout=5000)
            sav = dlg.locator(".el-dialog__footer button:has-text('保存')").first
            if sav.count() == 0:
                sav = dlg.locator("button:has-text('保存')").first
            sav.click(timeout=5000)
            try:
                dlg.wait_for(state="hidden", timeout=12000)
            except Exception:
                time.sleep(0.8)
            time.sleep(0.45)
            return True

        def _read_feed_state(feed_cell):
            feed_input = feed_cell.locator(f"{feed_ph}, input.el-input__inner").first
            feed_val = ""
            if feed_input.count() > 0:
                try:
                    feed_val = feed_input.input_value().strip()
                except Exception:
                    feed_val = (feed_input.get_attribute("value") or "").strip()
                if not feed_val:
                    feed_val = (feed_input.get_attribute("value") or "").strip()
                if not feed_val:
                    feed_val = (feed_input.get_attribute("placeholder") or "").strip()
            cell_text = feed_cell.inner_text().strip()
            cell_text = " ".join(cell_text.split())
            return feed_val, cell_text

        def _投料工位仍为空(feed_val, cell_text):
            ph_tokens = ("选择投料工位", "请选择投料工位")
            if feed_val and feed_val not in ("请选择",) + ph_tokens:
                return False
            invalid_text = ("", "请选择", "选择投料工位", "请选择投料工位", "-", "--", "删除")
            if cell_text in invalid_text:
                return True
            if any(t in cell_text for t in ph_tokens):
                return True
            return False

        def _校验成品半成品主料均为是():
            bad = []
            rs = page.locator("table.el-table__body tr")
            for j in range(rs.count()):
                r = rs.nth(j)
                cs = r.locator("td")
                if cs.count() < 5:
                    continue
                mk = cs.nth(3).inner_text().strip()
                if ("成品" not in mk) and ("半成品" not in mk):
                    continue
                major_col = cs.nth(4)
                yes_radio = major_col.locator("label.el-radio").filter(has_text="是").first
                ok = (
                    yes_radio.locator("span.el-radio__input.is-checked").count() > 0
                    if yes_radio.count() > 0
                    else False
                )
                if not ok:
                    bad.append((j + 1, cs.nth(2).inner_text().strip(), mk))
            if bad:
                detail = "，".join([f"第{a}行({b}/{c})" for a, b, c in bad])
                print(f"⚠️ 成品/半成品「是否主料」未选「是」：{detail}")
            else:
                print("✅ 成品/半成品「是否主料」均为「是」")

        missing_feed = 0
        for i in range(row_count):
            row = rows.nth(i)
            cells = row.locator("td")
            cell_count = cells.count()
            if cell_count < 5:
                continue

            manufacture_way = cells.nth(2).inner_text().strip()
            material_kind = cells.nth(3).inner_text().strip()
            major_col = cells.nth(4)
            feed_cell = _resolve_feed_cell(row, cells, cell_count)

            if ("成品" in material_kind) or ("半成品" in material_kind):
                yes_radio = major_col.locator("label.el-radio").filter(has_text="是").first
                yes_checked = (
                    yes_radio.locator("span.el-radio__input.is-checked").count() > 0
                    if yes_radio.count() > 0
                    else False
                )
                if not yes_checked:
                    if yes_radio.count() > 0:
                        yes_radio.click(timeout=5000, force=True)
                    else:
                        major_col.locator("label.el-radio span.el-radio__inner").first.click(timeout=5000, force=True)
                    print(f"✅ 第{i + 1}行({manufacture_way}/{material_kind}) 主料已自动改为“是”")

            feed_val, cell_text = _read_feed_state(feed_cell)
            if not _投料工位仍为空(feed_val, cell_text):
                print(f"✅ 第{i + 1}行({manufacture_way}/{material_kind}) 投料工位显示值：{cell_text or feed_val}")
                continue

            print(f"🔧 第{i + 1}行({manufacture_way}/{material_kind}) 打开添加工位并选择「机包」…")
            if not _投料工位弹窗选机包并保存(feed_cell):
                missing_feed += 1
                print(f"⚠️ 第{i + 1}行({manufacture_way}/{material_kind}) 投料工位未设置成功")
                continue

            feed_val2, cell_text2 = _read_feed_state(feed_cell)
            if _投料工位仍为空(feed_val2, cell_text2):
                missing_feed += 1
                print(f"⚠️ 第{i + 1}行({manufacture_way}/{material_kind}) 保存后投料工位仍为空")
            else:
                print(f"✅ 第{i + 1}行({manufacture_way}/{material_kind}) 已选投料工位：{cell_text2 or feed_val2}")

        if missing_feed == 0:
            print("✅ 投料工位检查通过：全部已选择")
        else:
            print(f"⚠️ 投料工位检查完成：有 {missing_feed} 行未选择")

        _校验成品半成品主料均为是()

    def _编辑页确认并提交():
        """
        编辑页可能加载较慢，或按钮在底部固定栏内；仅 wait_for 单一 locator 易超时。
        """
        try:
            page.locator(
                "button:has-text('添加物料'), button:has-text('+ 添加物料'), button:has-text('确认并提交')"
            ).first.wait_for(state="visible", timeout=18000)
        except Exception:
            pass

        sub = None
        deadline = time.time() + 28
        while time.time() < deadline:
            for loc in (
                page.locator("button.el-button--primary").filter(has_text="确认并提交").first,
                page.locator("button:has-text('确认并提交')").first,
                page.locator(".el-footer button").filter(has_text="确认并提交").first,
                page.locator(".fixed-footer button").filter(has_text="确认并提交").first,
                page.locator("footer button").filter(has_text="确认并提交").first,
            ):
                if loc.count() == 0:
                    continue
                try:
                    if loc.is_visible():
                        sub = loc
                        break
                except Exception:
                    continue
            if sub is not None:
                break
            time.sleep(0.35)

        if sub is None:
            raise RuntimeError(
                "未找到可点击的「确认并提交」：请确认已打开 BOM/加工产品编辑页（含「添加物料」或底部提交按钮），且未被弹窗遮挡。"
            )
        try:
            sub.scroll_into_view_if_needed(timeout=5000)
        except Exception:
            pass
        try:
            sub.click(timeout=8000, force=True)
        except Exception:
            sub.evaluate("el => el.click()")
        time.sleep(0.75)
        for loc in (
            page.locator(".el-message-box__wrapper:visible button.el-button--primary:has-text('确定')").first,
            page.locator(".el-message-box__wrapper:visible button:has-text('确定')").first,
            page.locator(".el-dialog:visible button:has-text('确定')").first,
        ):
            if loc.count() > 0:
                try:
                    if loc.is_visible():
                        loc.click(timeout=4000)
                        time.sleep(0.55)
                except Exception:
                    pass
        try:
            page.wait_for_selector("button:has-text('确认并提交')", state="hidden", timeout=25000)
        except Exception:
            time.sleep(1.2)

    def _待发布勾选首行并发布():
        rows_loc = _bom主表数据行()
        if rows_loc.count() == 0:
            print("⚠️ 待发布列表无数据行，跳过「发布」")
            return
        r = rows_loc.first
        r.scroll_into_view_if_needed(timeout=8000)
        checked = False
        if _按容器xpath点列表行勾选(1):
            inp = r.locator("td").first.locator("input.el-checkbox__original").first
            if inp.count() > 0:
                try:
                    checked = inp.evaluate("el => !!el.checked")
                except Exception:
                    checked = r.locator("label.el-checkbox.is-checked").count() > 0
            else:
                checked = r.locator("label.el-checkbox.is-checked").count() > 0
        if not checked:
            cb = r.locator("label.el-checkbox").first
            if cb.count() > 0:
                try:
                    cb.click(timeout=5000, force=True)
                except Exception:
                    cb.evaluate("el => el.click()")
        time.sleep(0.35)
        pub = page.locator("button:has-text('发布')").first
        if pub.count() == 0:
            print("⚠️ 未找到「发布」按钮")
            return
        try:
            pub.click(timeout=8000)
        except Exception:
            pub.click(timeout=8000, force=True)
        time.sleep(0.55)
        ok_btn = page.locator(".el-message-box__wrapper:visible button:has-text('确定')").first
        if ok_btn.count() > 0:
            try:
                if ok_btn.is_visible():
                    ok_btn.click(timeout=5000)
                    time.sleep(0.65)
            except Exception:
                pass

    def _提交编辑后回到列表并发布当前代码(code):
        print("📤 确认并提交…")
        _编辑页确认并提交()
        print("📋 回到待发布并查询当前代码后发布…")
        page.click("div.el-tabs__item:text('待发布')", timeout=8000)
        time.sleep(0.55)
        bom_code_input.click(timeout=5000)
        bom_code_input.fill(code)
        page.click("button:has-text('查询')", timeout=5000)
        time.sleep(1.2)
        _待发布勾选首行并发布()
        print("✅ 已尝试发布当前物料")

    def _待审核表格数据行():
        """待审核页 el-table 行：仅用 table.el-table__body tr 在部分页签下 count 恒为 0，改用 body-wrapper / #container。"""
        for sel in (
            page.locator("#container .el-table__body-wrapper tbody tr"),
            page.locator("#container table.el-table__body tbody tr"),
            page.locator(".el-table__body-wrapper:visible tbody tr"),
            page.locator("table.el-table__body tbody tr"),
            page.locator("table.el-table__body tr"),
        ):
            try:
                if sel.count() > 0:
                    return sel
            except Exception:
                continue
        return page.locator("#container .el-table__body-wrapper tbody tr")

    def _待审核列表行():
        r = _bom主表数据行()
        if r.count() > 0:
            return r
        return _待审核表格数据行()

    def _待审核依次审核(excel_codes):
        """进入「待审核」后，按 Excel 从第一条代码起依次填入物料代码查询，再对结果审核（通过 + 备注 1）。"""
        print("📋 进入「待审核」，按 Excel 从第一条代码起查询并依次审核…")
        page.locator("div.el-tabs__item").filter(has_text="待审核").first.click(timeout=8000)
        time.sleep(0.55)

        audited = 0
        for aidx, code in enumerate(excel_codes, start=1):
            bom_code_input.click(timeout=5000)
            bom_code_input.fill(code)
            page.click("button:has-text('查询')", timeout=5000)
            print(f"🔎 待审核查询（{aidx}/{len(excel_codes)}）：{code}")
            time.sleep(1.0)

            rows_loc = _待审核列表行()
            for _ in range(35):
                if rows_loc.count() > 0:
                    break
                time.sleep(0.15)
                rows_loc = _待审核列表行()

            if rows_loc.count() == 0:
                print(f"ℹ️ 待审核无该代码查询结果，跳过：{code}")
                continue

            row = rows_loc.first
            row.scroll_into_view_if_needed(timeout=8000)
            checked = False
            if _按容器xpath点列表行勾选(1):
                inp = row.locator("td").first.locator("input.el-checkbox__original").first
                if inp.count() > 0:
                    try:
                        checked = inp.evaluate("el => !!el.checked")
                    except Exception:
                        checked = row.locator("label.el-checkbox.is-checked").count() > 0
                else:
                    checked = row.locator("label.el-checkbox.is-checked").count() > 0
            if not checked:
                cb = row.locator("label.el-checkbox").first
                if cb.count() > 0:
                    try:
                        cb.click(timeout=5000, force=True)
                    except Exception:
                        cb.evaluate("el => el.click()")
            time.sleep(0.35)
            audit_btn = page.locator("button:has-text('审核')").first
            if audit_btn.count() == 0:
                print(f"⚠️ 代码 {code}：未找到「审核」按钮，中止待审核流程")
                break
            try:
                audit_btn.click(timeout=8000)
            except Exception as e:
                print(f"⚠️ 代码 {code}：点击「审核」失败：{e}")
                break
            dlg = page.locator(".el-dialog:visible").filter(has_text="处理意见").first
            if dlg.count() == 0:
                dlg = page.locator(".el-dialog:visible").filter(has_text="BOM审核").first
            dlg.wait_for(state="visible", timeout=12000)
            sel_wrap = dlg.locator(".el-form-item").filter(has_text="处理意见").locator(".el-select").first
            if sel_wrap.count() == 0:
                sel_wrap = dlg.locator(".el-select").first
            sel_wrap.locator(".el-input__inner").first.click(timeout=5000)
            time.sleep(0.35)
            page.locator(".el-select-dropdown:visible .el-select-dropdown__item").filter(has_text="通过").first.click(
                timeout=5000
            )
            time.sleep(0.25)
            remark = dlg.locator(".el-form-item").filter(has_text="备注").locator("textarea").first
            if remark.count() == 0:
                remark = dlg.locator("textarea").first
            remark.fill("1")
            dlg.locator("button:has-text('提交')").first.click(timeout=8000)
            try:
                dlg.wait_for(state="hidden", timeout=15000)
            except Exception:
                time.sleep(0.9)
            time.sleep(0.55)
            audited += 1
            print(f"✅ 待审核已提交（{code}）第 {audited} 条（处理意见：通过，备注：1）")

        if audited == 0:
            tail = _待审核列表行().count()
            print(
                f"ℹ️ 待审核未成功提交任何一条（最后一次查询后表格行数 {tail}）。"
                f"请确认「待审核」页签下物料代码与 Excel 一致。"
            )
        else:
            print(f"✅ 待审核流程结束，共提交 {audited} 条")

    def _工艺管理页表格数据行():
        """工艺管理列表主表体行（与待审核页类似）。"""
        for sel in (
            page.locator("#container .el-table__inner-wrapper > .el-table__body-wrapper tbody tr"),
            page.locator("#container .el-table__body-wrapper tbody tr"),
            page.locator("#container table.el-table__body tbody tr"),
            page.locator(".el-table__body-wrapper:visible tbody tr"),
        ):
            try:
                if sel.count() > 0:
                    return sel
            except Exception:
                continue
        return page.locator("#container .el-table__body-wrapper tbody tr")

    def _关闭确认信息弹窗():
        """审核/发布后常弹出 MessageBox，会拦截后续点击。"""
        box = page.locator(".el-message-box__wrapper:visible").first
        if box.count() == 0:
            return
        for label in ("确定", "确认", "知道了", "关闭"):
            b = box.locator("button").filter(has_text=label).first
            if b.count() > 0:
                try:
                    b.click(timeout=4000)
                    time.sleep(0.4)
                    return
                except Exception:
                    pass
        btns = box.locator(".el-message-box__btns button")
        if btns.count() > 0:
            try:
                btns.first.click(timeout=4000)
                time.sleep(0.4)
            except Exception:
                pass

    def _解析工艺筛选输入框():
        """
        工艺管理「列表」筛选用的工艺编号输入框。
        必须优先匹配占位符含「工艺编号」，否则详情页「工艺编号」「工艺名称」等表单项
        也会含「工艺」，且常为 disabled，会抢走首个匹配导致 fill 永久超时。
        进一步：优先取「含查询按钮」的表单区域内的工艺编号框，避免与详情区同名占位冲突。
        """
        shells = (
            page.locator("#container .el-form").filter(has=page.locator("button:has-text('查询')")),
            page.locator("#container"),
        )
        for shell in shells:
            if shell.count() == 0:
                continue
            loc = shell.first.locator("input.el-input__inner[placeholder*='工艺编号']").first
            if loc.count() > 0:
                return loc
            loc = shell.first.locator("input.el-input__inner[placeholder*='请输入工艺编']").first
            if loc.count() > 0:
                return loc
        items = page.locator("#container .el-form-item").filter(has_text="工艺")
        for i in range(items.count()):
            it = items.nth(i)
            inp = it.locator("input.el-input__inner").first
            if inp.count() == 0:
                continue
            try:
                ph = (inp.get_attribute("placeholder") or "").strip()
            except Exception:
                ph = ""
            if "物料代码" in ph:
                continue
            if "工艺编号" in ph or "请输入工艺" in ph:
                return inp
            if "工艺名称" in ph:
                continue
        for i in range(items.count()):
            it = items.nth(i)
            inp = it.locator("input.el-input__inner").first
            if inp.count() == 0:
                continue
            try:
                ph = inp.get_attribute("placeholder") or ""
            except Exception:
                ph = ""
            if "物料代码" in ph or "工艺名称" in ph:
                continue
            return inp
        alt = page.locator("#container input.el-input__inner[placeholder*='工艺']").first
        if alt.count() > 0:
            return alt
        return None

    def _等待工艺管理页就绪(timeout_ms: int = 20000) -> bool:
        deadline = time.time() + timeout_ms / 1000.0
        while time.time() < deadline:
            _关闭确认信息弹窗()
            if _解析工艺筛选输入框() is not None:
                return True
            time.sleep(0.25)
        return False

    def _工艺搜索输入框():
        _关闭确认信息弹窗()
        loc = _解析工艺筛选输入框()
        if loc is not None:
            return loc
        raise RuntimeError(
            "未找到工艺管理页的「工艺」筛选框（可能仍停留在「产品BOM」或存在未关闭的确认弹窗）。"
        )

    def _工艺筛选框点击聚焦(proc_input) -> None:
        """部分页面筛选框需先点击获得焦点后才从 disabled 变为可输入。"""
        try:
            proc_input.scroll_into_view_if_needed(timeout=6000)
        except Exception:
            pass
        try:
            proc_input.click(timeout=5000)
        except Exception:
            try:
                proc_input.click(timeout=5000, force=True)
            except Exception:
                pass
        time.sleep(0.18)

    def _等待工艺搜索输入可编辑(timeout_ms: int = 24000) -> bool:
        """
        等待列表筛选框出现在 DOM 并点击聚焦。
        工艺编号框可能短时仍为 disabled，写入由 _工艺筛选框写入文本（含 evaluate）兜底，此处不强制 is_enabled。
        """
        deadline = time.time() + timeout_ms / 1000.0
        while time.time() < deadline:
            _关闭确认信息弹窗()
            try:
                inp = _工艺搜索输入框()
                if inp.count() > 0:
                    _工艺筛选框点击聚焦(inp)
                    return True
            except RuntimeError:
                pass
            time.sleep(0.22)
        return False

    def _工艺筛选框写入文本(proc_input, text: str) -> bool:
        """先点击聚焦再 fill；失败则去掉 disabled 并触发 input（适配 Vue / el-input）。"""
        v = (text or "").strip()
        for _ in range(3):
            _工艺筛选框点击聚焦(proc_input)
            try:
                proc_input.fill(v, timeout=12000)
                return True
            except Exception:
                time.sleep(0.2)
        try:
            proc_input.fill(v, timeout=12000)
            return True
        except Exception:
            pass
        try:
            proc_input.evaluate(
                """(el, val) => {
                    el.removeAttribute('disabled');
                    el.removeAttribute('readonly');
                    el.value = val;
                    el.dispatchEvent(new InputEvent('input', { bubbles: true }));
                    el.dispatchEvent(new Event('change', { bubbles: true }));
                }""",
                v,
            )
            return True
        except Exception:
            return False

    def _工艺管理查询按钮():
        btn = page.locator("#container").locator("button:has-text('查询')").first
        if btn.count() > 0:
            return btn
        return page.locator("button:has-text('查询')").first

    def _点击目标行的工艺编号(row, numeric_code: str) -> bool:
        """在已选定的目标行内点击工艺编号（编码列，通常为含数字代码的短文本或链接）。"""
        _表格行滚入可视区(row)
        nc = numeric_code.strip()
        tokens = [nc]
        if nc.isdigit():
            tokens.append(f"FL_{nc}")
        elif nc.upper().startswith("FL_"):
            suf = nc.split("_", 1)[-1]
            if suf.isdigit():
                tokens.append(suf)
        td_count = row.locator("td").count()
        scored = []
        for i in range(td_count):
            cell = row.locator("td").nth(i)
            try:
                txt = cell.inner_text().strip().replace("\n", " ")
            except Exception:
                continue
            if not any(t in txt for t in tokens) or "半成品" in txt:
                continue
            pure = "".join(txt.split())
            looks_like_code = bool(pure) and all(
                (ch.isascii() and (ch.isalnum() or ch in "_-.")) for ch in pure
            )
            scored.append((0 if looks_like_code else 1, len(txt), cell))
        scored.sort(key=lambda x: (x[0], x[1]))
        for _, __, best_cell in scored:
            link = best_cell.locator("a").first
            if link.count() > 0:
                try:
                    link.click(timeout=8000, force=True)
                    return True
                except Exception:
                    try:
                        link.evaluate(
                            "e => { e.scrollIntoView({ block: 'center', inline: 'nearest' }); e.click(); }"
                        )
                        return True
                    except Exception:
                        pass
            try:
                best_cell.click(timeout=8000, force=True)
                return True
            except Exception:
                try:
                    best_cell.evaluate(
                        "e => { e.scrollIntoView({ block: 'center', inline: 'nearest' }); e.click(); }"
                    )
                    return True
                except Exception:
                    pass
        for tok in tokens:
            link = row.locator("td a").filter(has_text=tok).first
            if link.count() > 0:
                try:
                    link.click(timeout=8000, force=True)
                    return True
                except Exception:
                    try:
                        link.evaluate("e => { e.scrollIntoView({ block: 'center' }); e.click(); }")
                        return True
                    except Exception:
                        pass
        return False

    def _展示值匹配期望(实际: str, 期望: str) -> bool:
        """下拉展示可能被截断，允许「期望全文包含于实际」「实际为期望前缀」等宽松匹配。"""
        a = (实际 or "").strip()
        e = (期望 or "").strip()
        if not e:
            return True
        if a == e:
            return True
        if e in a or a in e:
            return True
        ca, ce = "".join(a.split()), "".join(e.split())
        if ce in ca:
            return True
        if len(ce) >= 12 and ca.startswith(ce[:12]):
            return True
        return False

    def _工序记录报表严格匹配期望(实际: str, 期望: str) -> bool:
        """
        「工序记录报表」须与期望全文一致（仅折叠空白），不得使用 _展示值匹配期望 的宽松规则：
        否则 RC-9020-01-10 与 RC-9020-01-11 会因共用前缀 RC-9020-01-1 被误判为通过。
        """
        a = "".join((实际 or "").split())
        e = "".join((期望 or "").split())
        if not e:
            return not bool(a)
        return a == e

    def _取表单项按标签(label: str):
        """
        优先 label 文案；失败时用本地保存页面对照的稳定类名（#/produce/craft/modeling/setMaintain …）。
        例如：batchRecordReportKey / instructReportKey / instructParams（ElementUI el-form-item 附加 class）。
        """
        lb = (label or "").strip()
        vue_field_class = {
            "批记录报表": "batchRecordReportKey",
            "指令报表": "instructReportKey",
            "指令参数记录": "instructParams",
        }.get(lb)
        if vue_field_class:
            by_cls = page.locator(f".el-form-item.{vue_field_class}").first
            if by_cls.count() > 0:
                return by_cls
        loc = page.locator(".el-form-item").filter(
            has=page.locator(".el-form-item__label", has_text=label)
        ).first
        if loc.count() > 0:
            return loc
        loc = page.locator(".el-form-item").filter(has=page.locator(f"label:text('{label}')")).first
        if loc.count() > 0:
            return loc
        return page.locator(".el-form-item").filter(has_text=label).first

    def _读取表单项下拉展示(form_item) -> str:
        if form_item.count() == 0:
            return ""
        sel_in = form_item.locator(".el-select .el-input__inner").first
        if sel_in.count() > 0:
            try:
                v = sel_in.input_value()
                if v and str(v).strip():
                    return str(v).strip()
            except Exception:
                pass
            try:
                return sel_in.inner_text(timeout=3000).strip()
            except Exception:
                pass
        inp = form_item.locator("input.el-input__inner").first
        if inp.count() > 0:
            try:
                v = inp.input_value()
                if v and str(v).strip():
                    return str(v).strip()
            except Exception:
                pass
            try:
                return inp.inner_text(timeout=3000).strip()
            except Exception:
                pass
        return ""

    def _等待工艺详情表单就绪(timeout_ms: int = 15000) -> bool:
        deadline = time.time() + timeout_ms / 1000.0
        while time.time() < deadline:
            _关闭确认信息弹窗()
            it = _取表单项按标签("批记录报表")
            if it.count() > 0:
                t = _读取表单项下拉展示(it)
                if t or it.locator(".el-select").count() > 0:
                    return True
            if page.locator("text=工艺基本信息").first.count() > 0:
                try:
                    page.locator("text=工艺基本信息").first.wait_for(state="visible", timeout=800)
                    return True
                except Exception:
                    pass
            time.sleep(0.25)
        return False

    def _工艺详情三项报表期望值():
        return (
            ("批记录报表", "RC-9000-05-02 批包装记录（2025.10.26）"),
            ("指令报表", "RC-9001-15-03 中药饮片批包装指令（2025.10.26）"),
            ("指令参数记录", "饮片包装批指令-新"),
        )

    def _读取工序列表项文案(li_el):
        """li.working-procedure-item：<b class=\"index\"> 序号 + <div> 工序名。"""
        try:
            idx_n = li_el.locator("b.index").first
            div_n = li_el.locator("div").first
            idx_t = idx_n.inner_text().strip() if idx_n.count() > 0 else ""
            name_t = div_n.inner_text().strip() if div_n.count() > 0 else ""
            if idx_t and name_t:
                return f"{idx_t}、{name_t}"
            raw = li_el.inner_text().strip().replace("\n", " ")
            return raw or name_t or idx_t
        except Exception:
            return ""

    def _报表下拉括号统一(s: str) -> str:
        """选项 DOM 可能为半角括号，期望文案为全角「（」。"""
        return (s or "").replace("(", "（").replace(")", "）")

    def _表单项_el_select_点击选项节点(opt) -> bool:
        """部分页面选项文案在 li > span 内，点 span 更贴近手工操作。"""
        try:
            opt.scroll_into_view_if_needed(timeout=5000)
        except Exception:
            pass
        sp = opt.locator(":scope > span").first
        if sp.count() == 0:
            sp = opt.locator("span").first
        if sp.count() > 0:
            try:
                sp.click(timeout=7000)
                return True
            except Exception:
                try:
                    sp.evaluate("e => e.click()")
                    return True
                except Exception:
                    pass
        try:
            opt.click(timeout=7000)
            return True
        except Exception:
            try:
                opt.evaluate("e => e.click()")
                return True
            except Exception:
                return False

    def _展开表单项_el_select(wrap, label_hint: str = "") -> bool:
        """
        ElementUI el-select 的内层 input 常为 readonly（placeholder 如「请选择批记录报表」），
        不能只依赖普通 click；依次点 input / .el-input 外壳 / .el-select，并 force + evaluate 兜底。
        """
        inner = wrap.locator(".el-input__inner").first
        shell = wrap.locator(".el-input").first
        for target, force in (
            (inner, False),
            (inner, True),
            (shell, False),
            (shell, True),
            (wrap, False),
            (wrap, True),
        ):
            if target.count() == 0:
                continue
            try:
                target.scroll_into_view_if_needed(timeout=6000)
                target.click(timeout=7000, force=force)
                return True
            except Exception:
                continue
        try:
            if inner.count() > 0:
                inner.evaluate(
                    "e => { e.scrollIntoView({ block: 'center' }); e.dispatchEvent(new MouseEvent('click', { bubbles: true })); }"
                )
                return True
        except Exception:
            pass
        if label_hint:
            print(f"⚠️ 下拉「{label_hint}」：展开 el-select（readonly）失败")
        return False

    def _表单项_el_select_选择(label: str, option_text: str) -> bool:
        """工艺基本信息区 ElementUI 下拉：展开后在可见下拉层中选中文案（按 RC 码等关键字匹配）。"""
        _关闭确认信息弹窗()
        item = _取表单项按标签(label)
        if item.count() == 0:
            print(f"⚠️ 下拉「{label}」：未找到表单项")
            return False
        wrap = item.locator(".el-select").first
        if wrap.count() == 0:
            print(f"⚠️ 下拉「{label}」：非 el-select")
            return False
        if not _展开表单项_el_select(wrap, label):
            return False
        time.sleep(0.45)
        dd = page.locator(".el-select-dropdown:visible").last
        try:
            dd.wait_for(state="visible", timeout=12000)
        except Exception:
            print(f"⚠️ 下拉「{label}」：选项面板未在时限内出现")
            try:
                page.keyboard.press("Escape")
            except Exception:
                pass
            return False
        ot = _报表下拉括号统一((option_text or "").strip())
        token = ot.split()[0] if ot else ""
        if not token:
            try:
                page.keyboard.press("Escape")
            except Exception:
                pass
            return False
        want_modify = "修改" in ot
        items = dd.locator(".el-select-dropdown__item")
        exact_ok = []
        token_ok = []
        for i in range(min(items.count(), 200)):
            o = items.nth(i)
            try:
                if not o.is_visible():
                    continue
            except Exception:
                continue
            try:
                raw_t = o.inner_text().strip().replace("\n", " ")
            except Exception:
                continue
            t = _报表下拉括号统一(raw_t)
            if token not in t and token not in raw_t:
                continue
            if ("修改" in t or "修改" in raw_t) and not want_modify:
                continue
            if _展示值匹配期望(t, ot) or _展示值匹配期望(raw_t, ot):
                exact_ok.append(o)
            else:
                token_ok.append(o)
        pick = exact_ok[0] if exact_ok else (token_ok[0] if token_ok else None)
        if pick is None:
            opts = dd.locator(".el-select-dropdown__item").filter(has_text=token)
            if opts.count() > 0:
                pick = opts.first
        if pick is None:
            print(f"⚠️ 下拉「{label}」：无匹配选项（关键字 {token!r}）")
            try:
                page.keyboard.press("Escape")
            except Exception:
                pass
            return False
        if not _表单项_el_select_点击选项节点(pick):
            print(f"⚠️ 下拉「{label}」：点击选项失败")
            try:
                page.keyboard.press("Escape")
            except Exception:
                pass
            return False
        time.sleep(0.4)
        try:
            page.keyboard.press("Escape")
        except Exception:
            pass
        return True

    def _加工顺序确认删除弹窗():
        box = page.locator(".el-message-box__wrapper:visible").first
        if box.count() == 0:
            return
        for txt in ("确定", "确认", "删除", "是"):
            b = box.locator("button").filter(has_text=txt).first
            if b.count() > 0:
                try:
                    b.click(timeout=5000)
                    time.sleep(0.35)
                    return
                except Exception:
                    pass

    def _加工顺序删至保留包装工序(*, verbose: bool = False) -> bool:
        """删除工序直至仅剩 1 条；优先删掉不含「包装」的行（与校验逻辑一致）。"""
        _关闭确认信息弹窗()

        def _当前工序条数():
            w = page.locator("#container li.working-procedure-item")
            if w.count() == 0:
                w = page.locator("li.working-procedure-item")
            return w, w.count()

        _, n0 = _当前工序条数()
        if verbose and n0 > 0:
            print(f"      └─ 加工顺序当前共 {n0} 条")

        for _ in range(24):
            wp = page.locator("#container li.working-procedure-item")
            if wp.count() == 0:
                wp = page.locator("li.working-procedure-item")
            n = wp.count()
            if n <= 1:
                if verbose:
                    print(f"      └─ 加工顺序已剩 {max(n, 0)} 条，结束删除")
                return True
            del_i = None
            for i in range(n):
                tx = _读取工序列表项文案(wp.nth(i))
                if "包装" not in (tx or ""):
                    del_i = i
                    break
            if del_i is None:
                del_i = n - 1
            del_tx = _读取工序列表项文案(wp.nth(del_i)).strip() or f"第{del_i + 1}条"
            if verbose:
                print(f"      └─ 删除工序项（优先非「包装」）：{del_tx!r}")
            row = wp.nth(del_i)
            trash = row.locator(
                ".el-icon-delete, i[class*='el-icon-delete'], span[class*='delete'], "
                "[class*='icon-delete'], button:has-text('删除')"
            ).first
            if trash.count() == 0:
                ib = row.locator("i, svg, button")
                ic = ib.count()
                if ic > 0:
                    trash = ib.nth(ic - 1)
            try:
                trash.scroll_into_view_if_needed(timeout=5000)
                trash.click(timeout=6000)
            except Exception:
                try:
                    trash.evaluate("e => e.click()")
                except Exception as e:
                    print(f"⚠️ 加工顺序：点击删除失败：{e}")
                    return False
            time.sleep(0.45)
            _加工顺序确认删除弹窗()
            _关闭确认信息弹窗()
            time.sleep(0.4)
        return page.locator("li.working-procedure-item").count() <= 1

    def _尝试点击保存工艺相关按钮(*, verbose: bool = False) -> bool:
        for name in ("保存工艺信息", "保存工艺", "保存基本信息", "保存工序信息", "保存"):
            btn = page.locator("#container button.el-button").filter(has_text=name).first
            if btn.count() > 0:
                try:
                    if btn.is_visible():
                        btn.scroll_into_view_if_needed(timeout=5000)
                        btn.click(timeout=7000)
                        time.sleep(0.65)
                        _关闭确认信息弹窗()
                        if verbose:
                            print(f"      └─ 已点击「{name}」")
                        return True
                except Exception:
                    continue
        if verbose:
            print("      └─ 未找到可见的保存类按钮（保存工艺信息 / 保存工序信息 …）")
        return False

    def _尝试修复工艺详情页(code: str = ""):
        """若三项报表或加工顺序不符，在详情页改下拉、删多余工序并尝试保存。"""
        pfx = f"   [{code}] " if code else "   "
        print(f"{pfx}──────── 自动修正过程 ────────")
        print(f"{pfx}① 三项报表下拉（批记录 / 指令报表 / 指令参数记录）")
        for lab, exp in _工艺详情三项报表期望值():
            item = _取表单项按标签(lab)
            if item.count() == 0:
                print(f"{pfx}   ⚠️ 「{lab}」：未找到表单项，跳过")
                continue
            cur = _读取表单项下拉展示(item)
            if _展示值匹配期望(cur, exp):
                print(f"{pfx}   ○ 「{lab}」已是目标值，跳过（当前：{cur!r}）")
                continue
            print(f"{pfx}   ⋯ 「{lab}」当前 {cur!r} → 期望 {exp!r}")
            ok_sel = _表单项_el_select_选择(lab, exp)
            print(f"{pfx}   {'✅' if ok_sel else '❌'} 「{lab}」下拉选择{'完成' if ok_sel else '失败'}")
            time.sleep(0.25)
        print(f"{pfx}② 加工顺序（删至仅保留含「包装」的一条）")
        ok_del = _加工顺序删至保留包装工序(verbose=True)
        print(f"{pfx}   {'✅' if ok_del else '❌'} 加工顺序精简{'完成' if ok_del else '异常退出'}")
        time.sleep(0.35)
        print(f"{pfx}③ 保存工艺/工序信息")
        ok_save = _尝试点击保存工艺相关按钮(verbose=True)
        print(f"{pfx}   {'✅' if ok_save else '⚠️'} 保存按钮{'已点击' if ok_save else '未点到（可能需手工保存）'}")
        time.sleep(0.9)
        print(f"{pfx}──────── 自动修正结束（即将复检） ────────")

    def _校验工艺基本信息三项报表字段():
        """
        进入工艺详情（工艺基本信息）后校验三个下拉展示是否与期望一致。
        """
        checks = _工艺详情三项报表期望值()
        lines = []
        all_ok = True
        for lab, exp in checks:
            item = _取表单项按标签(lab)
            if item.count() == 0:
                all_ok = False
                lines.append(f"❌ 「{lab}」：未找到表单项")
                continue
            actual = _读取表单项下拉展示(item)
            ok = _展示值匹配期望(actual, exp)
            if not ok:
                all_ok = False
            mark = "✅" if ok else "❌"
            lines.append(f"{mark} 「{lab}」展示：{actual!r} — 期望：{exp!r}")
        return all_ok, lines

    def _侧边点击一级菜单(module_name: str, timeout_ms: int = 15000) -> bool:
        """一级侧边菜单（如 生产管理）：`b.quick-name` 在折叠/滚动后可能失效，做多选择器兜底。"""
        candidates = (
            page.locator("b.quick-name").filter(has_text=module_name).first,
            page.locator(".el-submenu__title").filter(has_text=module_name).first,
            page.locator(".el-menu-item").filter(has_text=module_name).first,
            page.locator(".sidebar-container .el-menu").get_by_text(module_name, exact=True).first,
            page.locator(".sidebar-container").get_by_text(module_name, exact=True).first,
            page.locator("[class*='sidebar']").get_by_text(module_name, exact=True).first,
            page.get_by_role("menuitem", name=module_name).first,
            page.locator(".nest-menu-title, .submenu-title, .menu-title").filter(has_text=module_name).first,
        )
        for loc in candidates:
            try:
                if loc.count() == 0:
                    continue
                loc.scroll_into_view_if_needed(timeout=8000)
                loc.click(timeout=timeout_ms, force=True)
                return True
            except Exception:
                try:
                    loc.evaluate(
                        "e => { e.scrollIntoView({block:'center'}); e.click(); }"
                    )
                    return True
                except Exception:
                    continue
        return False

    def _工艺侧栏展开并点击子菜单(sub_name: str) -> bool:
        """
        侧栏：展开「工艺」分组 → 点击子菜单（标准工序 / 工艺管理 等）。
        子项 DOM 常见为 div.child-menu-item.click-interactive > div.child-menu-name，应点击整行。
        不强制先点「生产管理」：若已在生产模块且工艺已展开，可直接点行。
        优先在固定 XPath 宿主内查找（你提供的侧栏子区）：/html/body/div[3]/div[1]/div[2]/div/div[2]/div[1]
        """
        _关闭确认信息弹窗()
        tmo = 15000
        sub = (sub_name or "").strip()
        if not sub:
            return False

        _SIDEBAR_CHILD_MENU_HOST_XPATH = (
            "/html/body/div[3]/div[1]/div[2]/div/div[2]/div[1]"
        )

        def _侧栏子菜单宿主():
            loc = page.locator(f"xpath={_SIDEBAR_CHILD_MENU_HOST_XPATH}").first
            if loc.count() == 0:
                return None
            return loc

        def _点击_child_menu_item_整行(roots) -> bool:
            """在多个根下查找 child-menu-item 整行（优先 XPath 宿主）。"""
            for root in roots:
                if root is None:
                    continue
                try:
                    if hasattr(root, "count") and root.count() == 0:
                        continue
                except Exception:
                    continue
                row = root.locator("div.child-menu-item.click-interactive").filter(
                    has=page.locator("div.child-menu-name", has_text=re.compile("^" + re.escape(sub) + r"\s*$"))
                ).first
                if row.count() == 0:
                    row = root.locator("div.child-menu-item.click-interactive").filter(
                        has=page.locator("div.child-menu-name", has_text=sub)
                    ).first
                if row.count() == 0:
                    row = root.locator("div.child-menu-item.click-interactive").filter(has_text=sub).first
                if row.count() == 0:
                    continue
                try:
                    row.scroll_into_view_if_needed(timeout=8000)
                    row.click(timeout=tmo, force=True)
                    time.sleep(0.75)
                    _关闭确认信息弹窗()
                    print(f"✅ 侧栏已点击「{sub}」（child-menu-item 整行）")
                    return True
                except Exception:
                    try:
                        row.evaluate(
                            "e => { e.scrollIntoView({block:'center'}); e.dispatchEvent(new MouseEvent('click',{bubbles:true,cancelable:true})); }"
                        )
                        time.sleep(0.75)
                        _关闭确认信息弹窗()
                        print(f"✅ 侧栏已触达「{sub}」（child-menu-item 脚本点击）")
                        return True
                    except Exception:
                        continue
            return False

        def _展开工艺分组(roots) -> None:
            for root in roots:
                if root is None:
                    continue
                try:
                    if hasattr(root, "count") and root.count() == 0:
                        continue
                except Exception:
                    continue
                try:
                    t = root.locator("span.title:text('工艺')").first
                    if t.count() > 0:
                        t.scroll_into_view_if_needed(timeout=8000)
                        t.click(timeout=tmo, force=True)
                        time.sleep(0.55)
                        return
                except Exception as e:
                    print(f"⚠️ 展开侧栏「工艺」分组失败：{e}")

        host = _侧栏子菜单宿主()
        roots_ordered = []
        if host is not None:
            roots_ordered.append(host)
        roots_ordered.append(page)

        # 1) 优先 XPath 宿主内整行，再全页
        if _点击_child_menu_item_整行(roots_ordered):
            return True
        _展开工艺分组(roots_ordered)
        if _点击_child_menu_item_整行(roots_ordered):
            return True

        # 2) 可选：点一级「生产管理」后再展开工艺（部分布局一级不叫 quick-name）
        if _侧边点击一级菜单("生产管理", timeout_ms=tmo):
            time.sleep(0.35)
        else:
            print("ℹ️ 侧栏未点到「生产管理」（可能已在该模块或一级入口不同），继续尝试工艺子菜单…")
        _展开工艺分组(roots_ordered)
        _关闭确认信息弹窗()
        if _点击_child_menu_item_整行(roots_ordered):
            return True

        # 3) 兜底：仅点 child-menu-name（旧逻辑），仍优先 XPath 宿主
        name_rows = []
        if host is not None:
            name_rows.append(host)
        name_rows.append(page)
        candidates = []
        for base in name_rows:
            candidates.extend(
                (
                    base.locator("div.child-menu-item.click-interactive div.child-menu-name").filter(
                        has_text=re.compile("^" + re.escape(sub) + r"\s*$")
                    ).first,
                    base.locator("div.child-menu-name").filter(has_text=sub).first,
                )
            )
        candidates.extend(
            (
                page.locator(".sidebar-container div.child-menu-name").filter(has_text=sub).first,
                page.locator(".nest-menu .child-menu-name").filter(has_text=sub).first,
                page.locator(".sidebar-container .el-menu-item").filter(has_text=sub).first,
                page.locator(".el-menu-item").filter(has_text=sub).first,
            )
        )
        for child in candidates:
            if child.count() == 0:
                continue
            try:
                child.scroll_into_view_if_needed(timeout=8000)
                child.click(timeout=tmo, force=True)
                time.sleep(0.75)
                _关闭确认信息弹窗()
                print(f"✅ 侧栏已点击「{sub}」")
                return True
            except Exception:
                try:
                    child.evaluate(
                        "e => { e.scrollIntoView({block:'center'}); e.dispatchEvent(new MouseEvent('click',{bubbles:true})); }"
                    )
                    time.sleep(0.75)
                    _关闭确认信息弹窗()
                    print(f"✅ 侧栏已触达「{sub}」（脚本点击）")
                    return True
                except Exception:
                    continue

        try:
            ok = page.evaluate(
                """(sub, hostXp) => {
                function firstByXPath(xp) {
                    try {
                        const r = document.evaluate(
                            xp, document, null,
                            XPathResult.FIRST_ORDERED_NODE_TYPE, null
                        );
                        return r.singleNodeValue;
                    } catch (e) { return null; }
                }
                const host = hostXp ? firstByXPath(hostXp) : null;
                const scope = (host && host.nodeType === 1) ? host : document.body;
                const rows = scope.querySelectorAll('div.child-menu-item.click-interactive');
                for (const row of rows) {
                    const nm = row.querySelector('div.child-menu-name');
                    const t = (nm ? nm.innerText : row.innerText || '').replace(/\\s+/g, ' ').trim();
                    if (t === sub || t.startsWith(sub)) {
                        row.dispatchEvent(new MouseEvent('click', { bubbles: true, cancelable: true }));
                        return true;
                    }
                }
                const root = document.querySelector('.sidebar-container')
                    || document.querySelector('.el-menu')
                    || document.body;
                const els = root.querySelectorAll('div.child-menu-name, a.el-menu-item, li.el-menu-item');
                for (const el of els) {
                    const t = (el.innerText || '').replace(/\\s+/g, ' ').trim();
                    if (!t) continue;
                    if (t === sub || t.startsWith(sub)) {
                        el.dispatchEvent(new MouseEvent('click', { bubbles: true, cancelable: true }));
                        return true;
                    }
                }
                return false;
            }""",
                sub,
                _SIDEBAR_CHILD_MENU_HOST_XPATH,
            )
            if ok:
                time.sleep(0.85)
                _关闭确认信息弹窗()
                print(f"✅ 侧栏「{sub}」已由页面脚本兜底点击（含 XPath 宿主）")
                return True
        except Exception as e:
            print(f"⚠️ 侧栏脚本兜底点击「{sub}」失败：{e}")

        print(f"⚠️ 未在侧栏点到「{sub}」（请核对菜单是否改名或需先展开权限）")
        return False

    def _进入标准工序() -> bool:
        """
        侧边栏「生产管理 → 工艺 → 标准工序」。
        在「工序记录报表」与期望不符并已退出工艺详情后，进入该页以便对照或维护标准工序数据。
        """
        print("📂 进入：工艺 -> 标准工序")
        if _工艺侧栏展开并点击子菜单("标准工序"):
            print("✅ 已进入「标准工序」")
            return True
        if _工艺侧栏展开并点击子菜单("标准工序管理"):
            print("✅ 已进入「标准工序管理」")
            return True
        return False

    def _详情页强制返回工艺管理列表() -> None:
        """不依赖「工艺」筛选框检测：多点几次「返回上一页」类按钮，再面包屑回列表，关遮罩。"""
        _关闭确认信息弹窗()
        rx = re.compile(r"(?:<\s*)?返回\s*上一页|返回上一页")
        backs = page.locator(
            "button, a, .el-button, .el-button--text, span"
        ).filter(has_text=rx)
        for i in range(min(backs.count(), 14)):
            el = backs.nth(i)
            try:
                el.click(timeout=4500, force=True)
                time.sleep(0.5)
            except Exception:
                try:
                    el.evaluate(
                        "e => { e.scrollIntoView({block:'center'}); e.click(); }"
                    )
                    time.sleep(0.5)
                except Exception:
                    continue
        _返回工艺管理列表页()
        time.sleep(0.45)
        _关闭确认信息弹窗()
        try:
            for _ in range(3):
                page.keyboard.press("Escape")
                time.sleep(0.15)
        except Exception:
            pass

    def _校验加工顺序并点击包装再验工序记录报表():
        """
        检查「加工顺序」是否仅一条且为包装（文案形如「1、包装」）；
        符合则点击该「包装」步骤，切到右侧「基本信息」并读取「工序记录报表」是否与期望一致。
        若工序记录报表与期望不符：点击「返回上一页」离开详情，并跳转侧边栏「工艺 → 标准工序」。
        返回 (all_ok, log_lines, navigated_to_standard)；navigated_to_standard 为 True 时表示已离开工艺详情并进入标准工序，
        调用方应跳过偏差修改/工艺升级等仍依赖详情或原列表的步骤。
        """
        expect_report = "RC-9020-01-11包装生产记录（2026.01.26）"
        lines = []
        all_ok = True
        # 步骤行须「整段」匹配且长度受限；支持「1、包装」「1 包装」「1包装」（序号与名称分列渲染时常见）
        pattern_step_line = re.compile(
            r"^\s*(?:"
            r"\d{1,4}\s*[\u3001、,，.\uFF0E]\s*\S+(?:\s+\S+){0,4}"
            r"|\d{1,4}\s+\S+(?:\s+\S+){0,4}"
            r"|\d{1,4}(?=[^\d\s])\S*(?:\s+\S+){0,4}"
            r")\s*$"
        )
        # 仅用于 Playwright locator 粗筛（后续仍用 pattern_step_line 校验 inner_text）
        pattern_step_hint = re.compile(
            r"\d{1,4}(?:\s*[\u3001、,，.\uFF0E]\s*\S|\s+\S+|[^\d\s]\S)"
        )

        def _是加工顺序步骤行文案(tx: str) -> bool:
            tx = (tx or "").strip().replace("\n", " ")
            if len(tx) > 64:
                return False
            return bool(pattern_step_line.match(tx))

        def _是工序项纯名称兜底(tx: str) -> bool:
            """仅用于 li.working-procedure-item：序号在 b.index 未读出时可能只剩「包装」等短名称。"""
            tx = (tx or "").strip().replace("\n", " ")
            if not tx or len(tx) > 16:
                return False
            return bool(re.fullmatch(r"[\u4e00-\u9fff]{1,12}", tx))

        root = page.locator("#container").first
        if root.count() == 0:
            root = page.locator("body")

        # 优先：Vue 工序列表项（用户提供 HTML：li.working-procedure-item + b.index + div）
        step_items_loc = None
        filtered = []
        wp = root.locator("li.working-procedure-item")
        if wp.count() == 0:
            wp = page.locator("li.working-procedure-item")
        if wp.count() > 0:
            texts = []
            for i in range(wp.count()):
                try:
                    tx = _读取工序列表项文案(wp.nth(i)).strip().replace("\n", " ")
                    if _是加工顺序步骤行文案(tx) or _是工序项纯名称兜底(tx):
                        texts.append(tx)
                except Exception:
                    pass
            if texts:
                filtered = texts
                step_items_loc = wp

        # 工艺详情页「加工顺序」列表：用户提供的路径（旧版 ul/span/li）
        if not filtered:
            for xp in (
                "//*[@id='container']/div/div/div/div[2]/div/div[2]/div[1]/div[3]/ul/span/li",
            ):
                loc = page.locator(f"xpath={xp}")
                if loc.count() == 0:
                    continue
                texts = []
                for i in range(loc.count()):
                    try:
                        tx = loc.nth(i).inner_text().strip().replace("\n", " ")
                        if _是加工顺序步骤行文案(tx):
                            texts.append(tx)
                    except Exception:
                        pass
                if texts:
                    filtered = texts
                    step_items_loc = loc
                    break

        def _加工顺序文案_eval():
            """在浏览器内扫文案：步骤列表常在侧栏，未必在 #container 内，故以 body 为主并在「加工顺序」附近兜底。"""
            try:
                return page.evaluate("""() => {
                    const lineRe = /^\\s*(?:\\d{1,4}\\s*[\\u3001、,，.\\uFF0E]\\s*\\S+(?:\\s+\\S+){0,4}|\\d{1,4}\\s+\\S+(?:\\s+\\S+){0,4}|\\d{1,4}(?=[^\\d\\s])\\S*(?:\\s+\\S+){0,4})\\s*$/;
                    const bareRe = /^[\\u4e00-\\u9fff]{1,12}$/;
                    function pushSeen(arr, seen, t) {
                        if (!t || t.length > 64) return;
                        const okNum = lineRe.test(t);
                        const okBare = bareRe.test(t) && t.length <= 16;
                        if (!okNum && !okBare) return;
                        if (seen.has(t)) return;
                        seen.add(t);
                        arr.push({ t, len: t.length });
                    }
                    function collectWorkingProcedure() {
                        const hits = [];
                        const seen = new Set();
                        document.querySelectorAll('li.working-procedure-item').forEach(li => {
                            const idx = li.querySelector('b.index');
                            const div = li.querySelector('div');
                            let t = '';
                            if (idx && div) {
                                const it = (idx.innerText || '').trim();
                                const nt = (div.innerText || '').trim();
                                if (it && nt) t = it + '、' + nt;
                            }
                            if (!t) t = (li.innerText || '').trim().replace(/\\s+/g, ' ');
                            else t = t.replace(/\\s+/g, ' ');
                            pushSeen(hits, seen, t);
                        });
                        return hits;
                    }
                    function scanRoot(rootEl, seen) {
                        const hits = [];
                        rootEl.querySelectorAll(
                            'div, span, li, p, a, button, td, .el-step__title'
                        ).forEach(el => {
                            const t = (el.innerText || '').trim().replace(/\\s+/g, ' ');
                            if (!lineRe.test(t) || t.length > 64) return;
                            if (seen.has(t)) return;
                            seen.add(t);
                            hits.push({ t, len: t.length });
                        });
                        return hits;
                    }
                    let hits = collectWorkingProcedure();
                    const seen = new Set(hits.map(h => h.t));
                    if (hits.length === 0) {
                        hits = scanRoot(document.body, seen);
                    }
                    if (hits.length === 0) {
                        const labels = [...document.body.querySelectorAll('div, span, label')].filter(
                            el => (el.innerText || '').trim() === '加工顺序'
                        );
                        for (const lab of labels) {
                            let p = lab;
                            for (let d = 0; d < 12 && p; d++, p = p.parentElement) {
                                hits = scanRoot(p, seen);
                                if (hits.length) break;
                            }
                            if (hits.length) break;
                        }
                    }
                    hits.sort((a, b) => a.len - b.len);
                    const out = [];
                    const seenOut = new Set();
                    for (const h of hits) {
                        if (seenOut.has(h.t)) continue;
                        seenOut.add(h.t);
                        out.push(h.t);
                    }
                    return out;
                }""")
            except Exception:
                return []

        if not filtered:
            filtered = _加工顺序文案_eval()
        if not filtered:
            broad = (
                "li, .el-menu-item, .el-tree-node__content, .el-timeline-item, "
                ".el-step__title, [role='menuitem'], [role='tab'], "
                "div[class*='step'], div[class*='order'], div[class*='sequence'], "
                "span[class*='title'], .cell"
            )
            raw_items = page.locator(broad).filter(has_text=pattern_step_hint)
            cnt = raw_items.count()
            step_texts = []
            for i in range(cnt):
                try:
                    step_texts.append(raw_items.nth(i).inner_text().strip().replace("\n", " "))
                except Exception:
                    step_texts.append("")
            filtered = [t for t in step_texts if _是加工顺序步骤行文案(t)]

        if not filtered:
            hdr = page.locator("text=加工顺序").first
            if hdr.count() > 0:
                for xp in (
                    "xpath=ancestor::div[contains(@class,'el-card')][1]",
                    "xpath=ancestor::div[contains(@class,'panel')][1]",
                    "xpath=ancestor::div[contains(@class,'aside')][1]",
                    "xpath=ancestor::div[position()<=14]",
                ):
                    panel = hdr.locator(xp)
                    if panel.count() == 0:
                        continue
                    lis = panel.first.locator("div, span, li").filter(has_text=pattern_step_hint)
                    for i in range(min(lis.count(), 40)):
                        try:
                            tx = lis.nth(i).inner_text().strip().replace("\n", " ")
                            if _是加工顺序步骤行文案(tx):
                                filtered.append(tx)
                        except Exception:
                            pass
                    if filtered:
                        break

        seen = set()
        uniq_texts = []
        for t in filtered:
            key = t[:48]
            if key in seen:
                continue
            seen.add(key)
            uniq_texts.append(t)
        filtered = uniq_texts

        click_pkg_ok = False
        if len(filtered) != 1:
            all_ok = False
            lines.append(f"❌ 加工顺序：期望仅 1 条（序号、名称），实际收集 {len(filtered)} 条：{filtered!r}")
        else:
            only = filtered[0]
            if "包装" not in only:
                all_ok = False
                lines.append(f"❌ 加工顺序：唯一项须含「包装」，实际：{only!r}")
            else:
                lines.append(f"✅ 加工顺序：仅 1 条且含包装 — {only!r}")
                click_pkg = None
                if step_items_loc is not None:
                    cand = step_items_loc.filter(has_text=only).first
                    if cand.count() > 0:
                        click_pkg = cand
                if click_pkg is None:
                    click_pkg = root.get_by_text(only, exact=True).first
                if click_pkg.count() == 0:
                    click_pkg = root.locator("li.working-procedure-item").filter(has_text="包装").first
                if click_pkg.count() == 0:
                    click_pkg = root.locator("li, .el-menu-item, .el-tree-node__content, div, span").filter(
                        has_text=re.compile(
                            r"\d{1,4}(?:\s*[\u3001、,，.\uFF0E]\s*包装|\s+包装|(?=[^\d\s])包装)"
                        )
                    ).first
                if click_pkg.count() == 0:
                    click_pkg = page.get_by_text(
                        re.compile(r"\d{1,4}(?:\s*[\u3001、,，.\uFF0E]\s*包装|\s+包装|(?=[^\d\s])包装)")
                    ).first
                if click_pkg.count() > 0:
                    try:
                        click_pkg.scroll_into_view_if_needed(timeout=6000)
                        click_pkg.click(timeout=6000)
                        time.sleep(0.55)
                        click_pkg_ok = True
                    except Exception:
                        try:
                            click_pkg.evaluate(
                                "e => { e.scrollIntoView({ block: 'center' }); e.click(); }"
                            )
                            time.sleep(0.55)
                            click_pkg_ok = True
                        except Exception as e:
                            all_ok = False
                            lines.append(f"❌ 点击加工顺序「包装」失败：{e}")
                else:
                    all_ok = False
                    lines.append("❌ 未定位到可点击的「包装」加工顺序项")

        if not click_pkg_ok:
            lines.append("ℹ️ 「工序记录报表」：未满足「仅一条包装」或未成功点击包装，跳过该项校验")
            return all_ok, lines, False

        lines.append("ℹ️ 已点击加工顺序「包装」，在「基本信息」中核对「工序记录报表」…")
        tab = page.locator(".el-tabs__item").filter(has_text="基本信息").first
        if tab.count() > 0:
            try:
                tab.click(timeout=5000)
                time.sleep(0.4)
            except Exception:
                pass

        time.sleep(0.35)
        item = _取表单项按标签("工序记录报表")
        if item.count() == 0:
            all_ok = False
            lines.append("❌ 「工序记录报表」：未找到表单项")
            return all_ok, lines, False

        actual = _读取表单项下拉展示(item)
        ok_r = _工序记录报表严格匹配期望(actual, expect_report)
        if not ok_r:
            all_ok = False
        mark = "✅" if ok_r else "❌"
        lines.append(f"{mark} 「工序记录报表」展示：{actual!r} — 期望：{expect_report!r}")
        if not ok_r:
            lines.append(
                "❌ 「工序记录报表」强行校验不通过：须与期望全文一致（已禁用宽松前缀/包含匹配）"
            )
        if ok_r:
            return all_ok, lines, False

        lines.append(
            "🔧 「工序记录报表」与期望不符：开始离开详情并进入「标准工序」"
        )
        _std_nav_total = 4

        def _std_nav_step(i: int, msg: str) -> None:
            lines.append(f"   ▸ 步骤 {i}/{_std_nav_total}：{msg}")

        _std_nav_step(1, "关闭可能遮挡的确认 / 提示弹窗")
        _关闭确认信息弹窗()
        _std_nav_step(
            2,
            "离开工艺详情：依次尝试「返回上一页」、面包屑「工艺管理」、浏览器后退与 Esc",
        )
        _详情页强制返回工艺管理列表()
        _std_nav_step(3, "侧栏导航：展开「工艺」并在子菜单中进入「标准工序」")
        ok_std = _进入标准工序()
        _std_nav_step(
            4,
            "侧栏跳转结果：已进入「标准工序」"
            if ok_std
            else "侧栏跳转结果：未确认进入「标准工序」（请在侧栏手工点「标准工序」）",
        )
        if not ok_std:
            print(
                "⚠️ 自动进入「标准工序」失败；请在侧栏依次点："
                "「生产管理」→「工艺」→「标准工序」（或「标准工序管理」）"
            )
        return all_ok, lines, True

    def _详情页尝试点击返回上一页() -> bool:
        """工艺详情顶栏常见「返回上一页」，优先于面包屑 / 浏览器后退。"""
        _关闭确认信息弹窗()
        candidates = (
            page.get_by_role("button", name="返回上一页"),
            page.locator("button").filter(has_text="返回上一页"),
            page.locator("a").filter(has_text="返回上一页"),
            page.locator(".el-button").filter(has_text="返回上一页"),
            page.locator("span").filter(has_text="返回上一页"),
        )
        for loc in candidates:
            try:
                if loc.count() == 0:
                    continue
                el = loc.first
                if not el.is_visible():
                    continue
                el.scroll_into_view_if_needed(timeout=5000)
                try:
                    el.click(timeout=6000)
                except Exception:
                    el.evaluate(
                        "e => { e.scrollIntoView({ block: 'center' }); e.click(); }"
                    )
                time.sleep(0.55)
                if _等待工艺管理页就绪(12000):
                    return True
            except Exception:
                continue
        return False

    def _进入产品bom待发布页签():
        """从其它模块回到「产品BOM」并切到「待发布」（与流程入口一致）。"""
        _关闭确认信息弹窗()
        page.click("b.quick-name:text('生产管理')", timeout=8000)
        time.sleep(0.35)
        page.click("span.title:text('工艺')", timeout=8000)
        time.sleep(0.35)
        page.click("div.child-menu-name:text('产品BOM')", timeout=8000)
        time.sleep(0.55)
        page.click("div.el-tabs__item:text('待发布')", timeout=8000)
        time.sleep(0.6)
        _关闭确认信息弹窗()

    def _bom待发布重头查询所有代码(excel_codes):
        """待发布页下按 Excel 顺序逐条填入物料代码并查询（不编辑）。"""
        for ridx, c in enumerate(excel_codes, start=1):
            _关闭确认信息弹窗()
            try:
                bom_code_input.click(timeout=5000)
            except Exception:
                bom_code_input.click(timeout=5000, force=True)
            bom_code_input.fill(c)
            try:
                page.click("button:has-text('查询')", timeout=5000)
            except Exception:
                page.click("button:has-text('查询')", timeout=5000, force=True)
            print(f"🔎 待发布重头查询（{ridx}/{len(excel_codes)}）：{c}")
            time.sleep(1.0)

    def _返回工艺管理列表页():
        """详情页点击工艺编号后需回到列表，否则下一轮无法填「工艺」查询。"""
        _关闭确认信息弹窗()
        scoped = page.locator(".el-breadcrumb").locator("a").filter(has_text="工艺管理")
        if scoped.count() > 0:
            crumb = scoped.last
        else:
            crumb = page.locator(".el-breadcrumb__item").filter(has_text="工艺管理").locator("a").first
        if crumb.count() > 0:
            try:
                crumb.click(timeout=6000)
                time.sleep(0.75)
                if _等待工艺管理页就绪(12000):
                    return
            except Exception:
                pass
        try:
            page.go_back(timeout=10000)
            time.sleep(0.65)
        except Exception:
            pass
        _等待工艺管理页就绪(12000)

    def _工艺管理点击列表工具栏按钮(button_text: str) -> bool:
        """工艺管理列表 `#container` 内横向按钮（如 偏差修改、工艺升级），多为 `el-button`。"""
        _关闭确认信息弹窗()
        scoped = page.locator("#container")
        bt = (button_text or "").strip()
        groups = (
            scoped.get_by_role("button", name=bt),
            scoped.locator("button.el-button").filter(has_text=bt),
            scoped.locator("button").filter(has_text=bt),
            scoped.locator(".el-button").filter(has_text=bt),
            page.get_by_role("button", name=bt),
            page.locator("button.el-button").filter(has_text=bt),
        )

        def _try_click(btn) -> bool:
            try:
                if not btn.is_visible():
                    return False
                try:
                    if not btn.is_enabled():
                        return False
                except Exception:
                    pass
                btn.scroll_into_view_if_needed(timeout=6000)
                btn.click(timeout=8000)
                return True
            except Exception:
                try:
                    btn.evaluate(
                        "e => { e.scrollIntoView({ block: 'center' }); e.click(); }"
                    )
                    return True
                except Exception:
                    return False

        for grp in groups:
            try:
                n = grp.count()
            except Exception:
                continue
            for i in range(n):
                btn = grp.nth(i)
                if not _try_click(btn):
                    continue
                saw_follow = False
                follow_selectors = (
                    ".el-message-box__wrapper",
                    ".el-dialog__wrapper",
                    ".el-overlay-dialog",
                    ".el-dialog",
                    "[role='dialog']",
                )
                deadline = time.time() + 14.0
                while time.time() < deadline:
                    try:
                        role_dlgs = page.get_by_role("dialog")
                        for j in range(min(role_dlgs.count(), 12)):
                            try:
                                if role_dlgs.nth(j).is_visible():
                                    saw_follow = True
                                    break
                            except Exception:
                                continue
                    except Exception:
                        pass
                    if saw_follow:
                        break
                    for sel in follow_selectors:
                        locs = page.locator(sel)
                        for j in range(min(locs.count(), 12)):
                            try:
                                if locs.nth(j).is_visible():
                                    saw_follow = True
                                    break
                            except Exception:
                                continue
                        if saw_follow:
                            break
                    if saw_follow:
                        break
                    time.sleep(0.12)
                if saw_follow:
                    print(f"✅ 已点击「{bt}」（已出现确认框或对话框，最长等待约 14s）")
                else:
                    print(
                        f"⚠️ 已派发「{bt}」点击，但在约 14s 内仍未检测到可见对话框；"
                        "若界面较慢或样式不同请人工点此按钮并确认"
                    )
                return True
        print(f"⚠️ 未找到可点击的「{bt}」按钮（可见且可用）")
        return False

    def _工艺管理点击偏差修改() -> bool:
        return _工艺管理点击列表工具栏按钮("偏差修改")

    def _工艺管理点击工艺升级() -> bool:
        return _工艺管理点击列表工具栏按钮("工艺升级")

    def _工艺列表工具栏操作后尝试点击确定(timeout_ms: int = 14000) -> None:
        """列表工具栏「偏差修改」「工艺升级」等操作后的 MessageBox / Dialog 底部「确定」。"""
        start = time.time()
        deadline = start + timeout_ms / 1000.0
        idle_rounds = 0

        def _任意可见对话框() -> bool:
            try:
                dlgs = page.get_by_role("dialog")
                for i in range(min(dlgs.count(), 14)):
                    try:
                        if dlgs.nth(i).is_visible():
                            return True
                    except Exception:
                        continue
            except Exception:
                pass
            try:
                if page.locator(".el-message-box__wrapper:visible").count() > 0:
                    return True
            except Exception:
                pass
            return False

        while time.time() < deadline:
            had_ui = _任意可见对话框()
            if not had_ui:
                idle_rounds += 1
                if time.time() - start > 1.25 and idle_rounds >= 18:
                    break
            else:
                idle_rounds = 0

            _关闭确认信息弹窗()
            wrappers = page.locator(".el-dialog__wrapper")
            for i in range(min(wrappers.count(), 10)):
                w = wrappers.nth(i)
                try:
                    if not w.is_visible():
                        continue
                except Exception:
                    continue
                btn = w.locator(".el-dialog__footer button.el-button--primary").first
                if btn.count() == 0:
                    btn = w.locator(".el-dialog__footer button").filter(has_text=re.compile("^(确定|确认)$")).first
                if btn.count() > 0:
                    try:
                        btn.click(timeout=5000)
                        time.sleep(0.55)
                        _关闭确认信息弹窗()
                        return
                    except Exception:
                        pass
            time.sleep(0.12)

    def _启用中页签():
        tab = page.locator("div.el-tabs__item").filter(has_text="启用中").first
        if tab.count() > 0:
            try:
                tab.click(timeout=5000)
                time.sleep(0.45)
            except Exception:
                pass

    def _工艺管理切换待发布页签() -> bool:
        """工艺管理子页签「待发布(…)」：与产品 BOM 的待发布不是同一套页签。"""
        _关闭确认信息弹窗()
        for shell in (page.locator("#container"), page):
            tab = shell.locator("div.el-tabs__item").filter(has_text="待发布").first
            if tab.count() == 0:
                continue
            try:
                if not tab.is_visible():
                    continue
                tab.scroll_into_view_if_needed(timeout=6000)
                tab.click(timeout=6000)
                time.sleep(0.65)
                return True
            except Exception:
                continue
        return False

    def _工艺管理待发布查询并打开编辑(code: str) -> bool:
        """
        偏差修改提交后工艺常落在「待发布」：需先切该页签、用工艺条件查询，再点行内「编辑」进入可改表单。
        """
        if not _工艺管理切换待发布页签():
            print(f"⚠️ 代码 {code}：未找到或未点开工艺管理「待发布」页签")
            return False
        _关闭确认信息弹窗()
        proc_input = _工艺搜索输入框()
        qbtn = _工艺管理查询按钮()
        if not _工艺筛选框写入文本(proc_input, code):
            try:
                proc_input.click(timeout=8000, force=True)
            except Exception:
                pass
            proc_input.fill(code)
        try:
            qbtn.click(timeout=8000)
        except Exception:
            _关闭确认信息弹窗()
            qbtn = _工艺管理查询按钮()
            qbtn.click(timeout=8000, force=True)
        time.sleep(1.0)
        rows_loc = _工艺管理页表格数据行()
        for _ in range(35):
            if rows_loc.count() > 0:
                break
            time.sleep(0.15)
            rows_loc = _工艺管理页表格数据行()
        if rows_loc.count() == 0:
            print(f"⚠️ 代码 {code}：工艺管理「待发布」下查询无结果，无法点编辑")
            return False
        target = None
        tidx = None
        n = rows_loc.count()
        for i in range(n):
            r = rows_loc.nth(i)
            if not _行包含半成品或bcp(r):
                target, tidx = r, i
                break
        if target is None or tidx is None:
            print(f"⚠️ 代码 {code}：待发布查询结果均为半成品/BCP，跳过编辑")
            return False
        if not _勾选并点击编辑(target, tidx):
            print(f"⚠️ 代码 {code}：待发布列表未能勾选并点击「编辑」")
            return False
        print(f"✅ 代码 {code}：已从工艺管理「待发布」进入编辑页")
        time.sleep(0.75)
        return True

    def _进入工艺管理():
        """
        审核结束后可能仍停留在「工艺 > 产品BOM」等页，`b.quick-name:生产管理` 不一定可见。
        顺序：可先关 MessageBox → 点「工艺管理」→ 展开「工艺」→ 兜底「生产管理」。
        以「工艺」筛选框出现为准（排除 BOM 的物料代码框），避免误以为已跳转。
        """
        print("📂 审核完成后进入：工艺 -> 工艺管理（按需展开侧边栏）")
        _关闭确认信息弹窗()
        tmo = 15000
        child = page.locator("div.child-menu-name:text('工艺管理')").first

        def _尝试点击工艺管理() -> bool:
            try:
                if child.count() == 0:
                    return False
                child.scroll_into_view_if_needed(timeout=8000)
                child.click(timeout=tmo, force=True)
                time.sleep(0.65)
                _关闭确认信息弹窗()
                _启用中页签()
                return _等待工艺管理页就绪(15000)
            except Exception:
                return False

        if _尝试点击工艺管理():
            return

        try:
            page.locator("span.title:text('工艺')").first.scroll_into_view_if_needed(timeout=8000)
            page.locator("span.title:text('工艺')").first.click(timeout=tmo)
            time.sleep(0.4)
        except Exception:
            pass

        if _尝试点击工艺管理():
            return

        if _侧边点击一级菜单("生产管理", timeout_ms=tmo):
            time.sleep(0.4)
        try:
            page.locator("span.title:text('工艺')").first.click(timeout=tmo)
            time.sleep(0.4)
        except Exception:
            pass
        _关闭确认信息弹窗()
        try:
            child.scroll_into_view_if_needed(timeout=8000)
            child.click(timeout=tmo, force=True)
        except Exception as e:
            raise RuntimeError("无法点击侧边菜单「工艺管理」") from e
        time.sleep(0.65)
        _关闭确认信息弹窗()
        _启用中页签()
        if _等待工艺管理页就绪(20000):
            return
        print("⚠️ 进入工艺管理未就绪，重试完整路径…")
        _关闭确认信息弹窗()
        _侧边点击一级菜单("生产管理")
        time.sleep(0.45)
        page.locator("span.title:text('工艺')").first.click(timeout=tmo)
        time.sleep(0.45)
        _关闭确认信息弹窗()
        child.scroll_into_view_if_needed(timeout=8000)
        child.click(timeout=tmo, force=True)
        time.sleep(0.65)
        _关闭确认信息弹窗()
        _启用中页签()
        if not _等待工艺管理页就绪(20000):
            raise RuntimeError(
                "未能进入「工艺管理」：关闭确认弹窗后仍未出现「工艺」筛选输入框（已排除「请输入物料代码」）。"
            )

    def _工艺管理详情修正并重验(code: str):
        """
        列表勾选后执行「偏差修改」或「工艺升级」等：等待详情 → 自动修正 → 复检。
        返回 (detail_ok, step_ok)。
        """
        print(
            f"   [{code}] (1/3) 等待工艺详情表单就绪"
            f"（操作后页面通常已在详情，无需再点工艺编号）…"
        )
        time.sleep(0.85)
        if not _等待工艺详情表单就绪(20000):
            print(
                f"⚠️ 代码 {code}：(1/3) 超时：未检测到详情表单；"
                f"若界面仍在列表请先手工进入该工艺详情后再跑本轮"
            )
            return False, False
        print(f"   [{code}]      ✅ 表单已就绪")
        print(f"   [{code}] (2/3) 修正报表三项 + 加工顺序 + 保存…")
        _尝试修复工艺详情页(code)
        print(f"   [{code}] (3/3) 修正后复检（三项报表 + 加工顺序 + 工序记录报表）…")
        if not _等待工艺详情表单就绪(18000):
            print(f"⚠️ 代码 {code}：(3/3) 修正动作后详情表单仍未就绪")
            return False, False
        ok3b, detail_lines_b = _校验工艺基本信息三项报表字段()
        for ln in detail_lines_b:
            print(ln)
        ok_step_b, step_lines_b, nav_std_b = _校验加工顺序并点击包装再验工序记录报表()
        for ln in step_lines_b:
            print(ln)
        if nav_std_b:
            return False, False
        return ok3b, ok_step_b

    def _工艺管理待发布编辑后修正并发布(code: str) -> None:
        """
        从工艺管理「待发布」列表进入编辑且表单已打开后：
        与详情路径一致做三项/加工顺序检查、自动修正与保存，再回到列表勾选当前查询行并点工具栏「发布」。
        """
        print(f"🔧 代码 {code}：待发布编辑页 — 检查、自动修正与保存复检…")
        detail_ok, step_ok = _工艺管理详情修正并重验(code)
        if detail_ok and step_ok:
            print(f"✅ 代码 {code}：待发布编辑页修正后校验通过")
        else:
            print(
                f"⚠️ 代码 {code}：待发布编辑页修正后仍有不符，"
                f"已尽量保存；将继续返回列表并尝试发布，请人工复核"
            )

        print(f"📤 代码 {code}：返回工艺管理列表并在「待发布」下尝试发布…")
        if not _详情页尝试点击返回上一页():
            _返回工艺管理列表页()
        else:
            time.sleep(0.35)
            _关闭确认信息弹窗()
            if not _等待工艺管理页就绪(12000):
                _返回工艺管理列表页()
        time.sleep(0.55)
        _关闭确认信息弹窗()

        if not _工艺管理切换待发布页签():
            print(f"⚠️ 代码 {code}：未能切回工艺管理「待发布」页签，跳过发布")
            return

        _关闭确认信息弹窗()
        proc_r = _工艺搜索输入框()
        qbtn_r = _工艺管理查询按钮()
        if not _工艺筛选框写入文本(proc_r, code):
            try:
                proc_r.click(timeout=8000, force=True)
            except Exception:
                pass
            proc_r.fill(code)
        try:
            qbtn_r.click(timeout=8000)
        except Exception:
            _关闭确认信息弹窗()
            qbtn_r = _工艺管理查询按钮()
            qbtn_r.click(timeout=8000, force=True)
        time.sleep(1.0)

        rows_r = _工艺管理页表格数据行()
        for _ in range(35):
            if rows_r.count() > 0:
                break
            time.sleep(0.15)
            rows_r = _工艺管理页表格数据行()
        if rows_r.count() == 0:
            print(f"⚠️ 代码 {code}：返回「待发布」后查询无行，无法发布")
            return

        if not _勾选表格行首列(rows_r.first, 0, use_bom_row_xpath=False):
            print(f"⚠️ 代码 {code}：发布前勾选失败")
            return

        time.sleep(0.35)
        if _工艺管理点击列表工具栏按钮("发布"):
            _工艺列表工具栏操作后尝试点击确定(16000)
            _关闭确认信息弹窗()
            print(f"✅ 代码 {code}：已尝试完成工艺管理「待发布」发布")
        else:
            print(
                f"⚠️ 代码 {code}：未点到可用的「发布」按钮"
                f"（可能需先保存、未勾选或按钮禁用）"
            )

    def _工艺管理依次搜索并点击工艺编号(excel_codes):
        """按 Excel 代码依次在「工艺」条件查询：先检视是否存在；不符则优先「偏差修改」修正，不行再「工艺升级」。"""
        _进入工艺管理()
        for gidx, code in enumerate(excel_codes, start=1):
            _关闭确认信息弹窗()
            proc_input = _工艺搜索输入框()
            qbtn = _工艺管理查询按钮()
            try:
                proc_input.click(timeout=8000)
            except Exception:
                proc_input.click(timeout=8000, force=True)
            proc_input.fill(code)
            try:
                qbtn.click(timeout=8000)
            except Exception:
                _关闭确认信息弹窗()
                qbtn = _工艺管理查询按钮()
                qbtn.click(timeout=8000, force=True)
            print(f"🔎 工艺管理查询（{gidx}/{len(excel_codes)}）：{code}")
            time.sleep(1.0)
            rows_loc = _工艺管理页表格数据行()
            for _ in range(35):
                if rows_loc.count() > 0:
                    break
                time.sleep(0.15)
                rows_loc = _工艺管理页表格数据行()
            if rows_loc.count() == 0:
                print(f"ℹ️ 工艺管理无查询结果（不存在该工艺条件），跳过：{code}")
                continue
            print(
                f"✅ 工艺管理存在查询结果：共 {rows_loc.count()} 行，继续检视（{code}）"
            )
            target = None
            target_row_index = None
            n = rows_loc.count()
            for i in range(n):
                r = rows_loc.nth(i)
                if not _行包含半成品或bcp(r):
                    target = r
                    target_row_index = i
                    break
            if target is None:
                print(f"⏭️ 工艺管理结果均为半成品/BCP，跳过：{code}")
                if gidx == len(excel_codes):
                    print(
                        f"📋 本轮最后一条（{gidx}/{len(excel_codes)}），"
                        f"跳转工艺管理「待发布」并按 {code} 查询…"
                    )
                    if not _工艺管理切换待发布页签():
                        print("⚠️ 未能切换到工艺管理「待发布」页签")
                    else:
                        _关闭确认信息弹窗()
                        proc_tp = _工艺搜索输入框()
                        qbtn_tp = _工艺管理查询按钮()
                        if not _工艺筛选框写入文本(proc_tp, code):
                            try:
                                proc_tp.click(timeout=8000, force=True)
                            except Exception:
                                pass
                            proc_tp.fill(code)
                        try:
                            qbtn_tp.click(timeout=8000)
                        except Exception:
                            _关闭确认信息弹窗()
                            qbtn_tp = _工艺管理查询按钮()
                            qbtn_tp.click(timeout=8000, force=True)
                        time.sleep(1.0)
                        print(f"✅ 已停留在「待发布」并完成查询：{code}")
                        rows_pub = _工艺管理页表格数据行()
                        for _ in range(35):
                            if rows_pub.count() > 0:
                                break
                            time.sleep(0.15)
                            rows_pub = _工艺管理页表格数据行()
                        if rows_pub.count() == 0:
                            print(
                                f"⚠️ 代码 {code}：「待发布」查询后无数据行，无法点编辑"
                            )
                        else:
                            tp_target = None
                            tp_idx = None
                            for i in range(rows_pub.count()):
                                rr = rows_pub.nth(i)
                                if not _行包含半成品或bcp(rr):
                                    tp_target, tp_idx = rr, i
                                    break
                            if tp_target is None:
                                tp_target, tp_idx = rows_pub.nth(0), 0
                            if _勾选并点击编辑(tp_target, tp_idx):
                                print(f"✅ 代码 {code}：已在「待发布」点击编辑进入表单")
                                time.sleep(0.75)
                                _工艺管理待发布编辑后修正并发布(code)
                            else:
                                print(
                                    f"⚠️ 代码 {code}：「待发布」列表未能勾选并点击「编辑」"
                                )
                continue
            if not _点击目标行的工艺编号(target, code):
                print(f"⚠️ 工艺管理：未能点击工艺编号（{code}）")
                continue
            print(f"✅ 已点击工艺编号（{code}，非半成品行）")
            time.sleep(0.85)
            detail_ok = True
            step_ok = True
            skip_deviation = False
            if not _等待工艺详情表单就绪(18000):
                print(f"⚠️ 代码 {code}：未检测到工艺详情表单（批记录报表），视为校验不符")
                detail_ok = False
                step_ok = False
            else:
                ok3, detail_lines = _校验工艺基本信息三项报表字段()
                for ln in detail_lines:
                    print(ln)
                detail_ok = ok3
                if ok3:
                    print(f"✅ 代码 {code}：批记录报表 / 指令报表 / 指令参数记录 均与期望一致（含截断宽松匹配）")
                else:
                    print(f"⚠️ 代码 {code}：上述三项与期望不一致或缺少表单项，请人工核对")
                ok_step, step_lines, skip_deviation = _校验加工顺序并点击包装再验工序记录报表()
                for ln in step_lines:
                    print(ln)
                step_ok = ok_step
                if skip_deviation:
                    step_ok = False
                if ok_step and not skip_deviation:
                    print(
                        f"✅ 代码 {code}：加工顺序（单条包装）及工序记录报表校验通过（含截断宽松匹配）"
                    )
                elif skip_deviation:
                    print(
                        f"❌ 代码 {code}：工序记录报表强行校验不通过，已尝试退出详情并进入「标准工序」"
                    )
                else:
                    print(f"⚠️ 代码 {code}：加工顺序或工序记录报表与期望不符，请人工核对")

            if skip_deviation:
                print(
                    f"ℹ️ 代码 {code}：因「工序记录报表」不符已退出详情并进入「标准工序」，"
                    f"跳过本轮自动「偏差修改 / 工艺升级」"
                )
                _进入工艺管理()
            elif not detail_ok or not step_ok:
                print(
                    f"🔧 代码 {code}：初检未通过；返回列表后"
                    f"先尝试「偏差修改」+ 自动修正，仍不符再「工艺升级」+ 自动修正"
                )
                if not _详情页尝试点击返回上一页():
                    _返回工艺管理列表页()
                else:
                    time.sleep(0.35)
                    _关闭确认信息弹窗()
                    if not _等待工艺管理页就绪(12000):
                        _返回工艺管理列表页()
                time.sleep(0.55)
                _关闭确认信息弹窗()
                rows_again = _工艺管理页表格数据行()
                for _ in range(35):
                    if rows_again.count() > 0:
                        break
                    time.sleep(0.15)
                    rows_again = _工艺管理页表格数据行()
                if target_row_index is None or rows_again.count() <= target_row_index:
                    print(
                        f"⚠️ 代码 {code}：返回列表后无法定位原数据行（期望行号 {target_row_index}），"
                        f"无法执行偏差修改 / 工艺升级"
                    )
                else:
                    pick = rows_again.nth(target_row_index)
                    if not _勾选表格行首列(pick, target_row_index, use_bom_row_xpath=False):
                        print(
                            f"⚠️ 代码 {code}：返回列表后勾选失败，请手动勾选后再执行偏差修改或工艺升级"
                        )
                    else:
                        tried_pc = False
                        print(f"🔧 代码 {code}：① 尝试「偏差修改」…")
                        if _工艺管理点击偏差修改():
                            tried_pc = True
                            _工艺列表工具栏操作后尝试点击确定(16000)
                            time.sleep(0.75)
                            _关闭确认信息弹窗()
                            print(
                                f"🔧 代码 {code}：「偏差修改」已执行，"
                                f"进入工艺管理「待发布」查询并点「编辑」后再自动修正与复检…"
                            )
                            if not _工艺管理待发布查询并打开编辑(code):
                                print(
                                    f"ℹ️ 代码 {code}：待发布编辑入口未打开，"
                                    f"仍在当前页尝试等待详情并复检（可能超时）…"
                                )
                            detail_ok, step_ok = _工艺管理详情修正并重验(code)
                            if detail_ok and step_ok:
                                print(f"✅ 代码 {code}：偏差修改并修正后校验通过")
                        else:
                            print(
                                f"ℹ️ 代码 {code}：未找到可点击的「偏差修改」"
                                f"（将尝试「工艺升级」）"
                            )

                        if not detail_ok or not step_ok:
                            why = "偏差修改后仍不符" if tried_pc else "偏差修改不可用或未改善"
                            print(f"🔧 代码 {code}：② {why}，尝试「工艺升级」…")
                            _返回工艺管理列表页()
                            time.sleep(0.55)
                            _关闭确认信息弹窗()
                            rows_b = _工艺管理页表格数据行()
                            for _ in range(35):
                                if rows_b.count() > 0:
                                    break
                                time.sleep(0.15)
                                rows_b = _工艺管理页表格数据行()
                            if (
                                target_row_index is not None
                                and rows_b.count() > target_row_index
                            ):
                                pick2 = rows_b.nth(target_row_index)
                                if _勾选表格行首列(
                                    pick2, target_row_index, use_bom_row_xpath=False
                                ):
                                    if _工艺管理点击工艺升级():
                                        _工艺列表工具栏操作后尝试点击确定(16000)
                                        time.sleep(0.75)
                                        _关闭确认信息弹窗()
                                        print(
                                            f"🔧 代码 {code}：「工艺升级」已执行，开始自动修正与复检…"
                                        )
                                        detail_ok, step_ok = _工艺管理详情修正并重验(code)
                                        if detail_ok and step_ok:
                                            print(
                                                f"✅ 代码 {code}：工艺升级并修正后校验通过"
                                            )
                                        else:
                                            print(
                                                f"⚠️ 代码 {code}：工艺升级并修正后仍有不符，请人工核对"
                                            )
                                    else:
                                        print(
                                            f"⚠️ 代码 {code}：未能点击「工艺升级」"
                                        )
                                else:
                                    print(
                                        f"⚠️ 代码 {code}：工艺升级前勾选失败"
                                    )
                            else:
                                print(
                                    f"⚠️ 代码 {code}：工艺升级前列表行不可用（期望行号 {target_row_index}）"
                                )

            if not skip_deviation:
                _返回工艺管理列表页()
                if gidx == len(excel_codes):
                    print(
                        "📋 工艺管理本轮最后一条代码已处理，跳转产品BOM「待发布」并按 Excel 重头查询…"
                    )
                    try:
                        _进入产品bom待发布页签()
                        _bom待发布重头查询所有代码(excel_codes)
                    except Exception as e:
                        print(f"⚠️ 跳转待发布并重头查询失败：{e}")

    processed_any = False
    for idx, code in enumerate(codes, start=1):
        # 搜索之前切换到「待发布」
        page.click("div.el-tabs__item:text('待发布')", timeout=8000)
        time.sleep(0.6)

        bom_code_input.click(timeout=5000)
        bom_code_input.fill(code)
        page.click("button:has-text('查询')", timeout=5000)
        print(f"🔎 查询（{idx}/{len(codes)}）：{code}")
        time.sleep(1.5)

        rows_loc = _bom主表数据行()
        row_count = rows_loc.count()
        if row_count == 0:
            print("无需发布 查询下一条")
            continue
        print(f"📌 查询结果行数：{row_count}（主表体，已排除固定列重复）")

        chosen_row_index = None
        if row_count == 1:
            row = _首行row_按xpath() or rows_loc.first
            chosen_row_index = 0
            first_name = _首行物料名称_按xpath()
            if (first_name and ("半成品" in first_name or "bcp" in first_name.lower())) or _行包含半成品或bcp(row):
                print("⏭️ 单行但物料名称含半成品/BCP，查询下一条")
                continue
        elif row_count == 2:
            first_row = _首行row_按xpath() or rows_loc.nth(0)
            second_row = rows_loc.nth(1)
            # 两行规则：首行含半成品 -> 选第二行；否则选第一行
            first_name = _首行物料名称_按xpath()
            first_is_half = (first_name and ("半成品" in first_name or "bcp" in first_name.lower())) or _行包含半成品或bcp(first_row)
            row = second_row if first_is_half else first_row
            chosen_row_index = 1 if first_is_half else 0
            if _行包含半成品或bcp(row):
                print("⏭️ 两行选择结果仍含半成品/BCP，查询下一条")
                continue
        else:
            # 兜底：多行时选第一条非半成品/BCP
            row = None
            for i in range(row_count):
                r = rows_loc.nth(i)
                if not _行包含半成品或bcp(r):
                    row = r
                    chosen_row_index = i
                    break
            if row is None:
                print("⏭️ 多行结果均为半成品/BCP，查询下一条")
                continue

        ok = _勾选并点击编辑(row, list_row_index=chosen_row_index)
        if not ok:
            raise RuntimeError("未找到可点击的编辑入口")
        print("✅ 已勾选并点击编辑")
        _进入编辑后检查投料工位与主料()
        _提交编辑后回到列表并发布当前代码(code)
        processed_any = True
        continue

    if not processed_any:
        print("⚠️ Excel 代码已遍历完，未找到可编辑的非半成品单行结果")
    else:
        print("✅ Excel 代码本轮编辑并发布流程已跑完")
    print("开始待审核：按 Excel 从第一条代码起查询…")
    _待审核依次审核(codes)
    print("待审核结束后进入工艺管理：按 Excel 依次查询并点击非半成品行的工艺编号…")
    _工艺管理依次搜索并点击工艺编号(codes)


def 更新问题(page):
    print("已选择：更新问题")
    print(page.title())


def main():
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=False)
        page = browser.new_page()
        page.goto("http://111.10.250.16:9803/#/main/home")
        page.wait_for_load_state("domcontentloaded")
        page.fill('input[placeholder="请输入用户名"]', "5894")
        page.fill('input[placeholder="请输入密码"]', "Cq123456")
        page.click("button:has-text('登录')")

        print("\n1 更新计划")
        print("2 更新问题")
        choice = input("请输入选项：").strip()

        if choice == "1":
            更新计划(page)
        elif choice == "2":
            更新问题(page)
        else:
            print("无效选项")

        # 流程结束后自动关闭浏览器（如需观察页面可在此调整等待时间）
        time.sleep(2)
        browser.close()


if __name__ == "__main__":
    main()
